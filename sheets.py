"""Google Sheets backend for HPCL SOD MIS."""

# ── SSL bypass for corporate networks with SSL-inspection proxies ─────────────
# Must run before any HTTPS-related imports.
import ssl
import urllib3

ssl._create_default_https_context = ssl._create_unverified_context
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    from requests.adapters import HTTPAdapter as _HA
    _orig_send = _HA.send
    def _send_no_ssl(self, request, **kw):
        kw["verify"] = False
        return _orig_send(self, request, **kw)
    _HA.send = _send_no_ssl
except Exception:
    pass
# ─────────────────────────────────────────────────────────────────────────────

import json
import time
from datetime import datetime, date, timedelta, timezone

import gspread
import streamlit as st
from google.oauth2.service_account import Credentials

import bcrypt
import psycopg2
import psycopg2.extras
import psycopg2.pool


# ── Postgres (staging branch: functions are being ported off Google Sheets
#    one domain at a time -- see the migration map artifact. Anything not
#    yet listed there still runs against Sheets via the code below.) ────────

@st.cache_resource
def _pg_pool():
    return psycopg2.pool.SimpleConnectionPool(
        1, 20, st.secrets["postgres"]["database_url"], sslmode="require"
    )


def _pg_query(sql: str, params: tuple = (), fetch: bool = True):
    """Run a query against Postgres. fetch=True returns list[dict]; fetch=False
    returns the affected row count. Always commits on success, rolls back on error."""
    pool = _pg_pool()
    conn = pool.getconn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            if fetch:
                rows = cur.fetchall()
                conn.commit()
                return [dict(r) for r in rows]
            conn.commit()
            return cur.rowcount
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


def _pg_one(sql: str, params: tuple = ()):
    """Run a query and return the first row as a dict, or None."""
    rows = _pg_query(sql, params, fetch=True)
    return rows[0] if rows else None


def _mk_to_date(month_year: str) -> date:
    """'Apr-2026' -> date(2026, 4, 1), for querying Postgres's date column.
    month_key(d) does the reverse conversion and already exists below."""
    month_0, year = parse_month_key(month_year)
    return date(year, month_0 + 1, 1)


def _api_call(fn, *args, retries: int = 6, **kwargs):
    """Execute a Sheets API call with exponential backoff on 429 quota errors
    and on transient connection failures (DNS blips, dropped connections --
    the same class of error as an occasional 'Failed to resolve
    sheets.googleapis.com' seen on flaky networks).
    Waits 2^n seconds between attempts (1 s, 2 s, 4 s, 8 s, 16 s, 32 s)."""
    import requests as _requests
    for n in range(retries):
        try:
            return fn(*args, **kwargs)
        except gspread.exceptions.APIError as exc:
            resp = getattr(exc, "response", None)
            code = getattr(resp, "status_code", 0) if resp else 0
            if code == 429 and n < retries - 1:
                time.sleep(min(2 ** n, 32))
            else:
                raise
        except (_requests.exceptions.ConnectionError, _requests.exceptions.Timeout):
            if n < retries - 1:
                time.sleep(min(2 ** n, 32))
            else:
                raise

_SCOPES = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive",
]

TABS = {
    "USER_ACCESS":        "UserAccess",
    "LOCATION_MASTER":    "LocationMaster",
    "HELPDESK":           "HelpDesk",
    "AUDIT_LOG":          "AuditLog",
    "MIS_DRAFT":          "MIS_DRAFT",
    "MIS_SUBMITTED":      "MIS_Submitted",
    "SUBMISSION_STATUS":  "SubmissionStatus",
    "RAILWAY_CLAIMS":     "Railway_Claims",
    "IRR_DETAILS":        "IRR_Details",
    "LEGAL_CASES":        "Legal_Cases",
    "REVISION_REQUESTS":    "RevisionRequests",
    "SETTINGS":             "Settings",
    # M&I MIS separate tabs
    "MI_TANK_OUTAGE":       "MI_TANK_OUTAGE",
    "MI_MAJOR_REPAIR":      "MI_MAJOR_REPAIR",
    "MI_VRU":               "MI_VRU",
    "MI_AUDIT_2526":        "MI_AUDIT_2526",
    "MI_AUDIT_2627":        "MI_AUDIT_2627",
    "MI_TECH_AUDIT":        "MI_TECH_AUDIT",
    "MI_EQUIP_BREAKDOWN":   "MI_EQUIP_BREAKDOWN",
    "MI_INT_PIPELINE":      "MI_INT_PIPELINE",
    "MI_EXT_PIPELINE":      "MI_EXT_PIPELINE",
    "MI_TANK_STATUS":       "MI_TANK_STATUS",
    "TANK_MASTER":          "TankMaster",
    "EMAIL_MASTER":         "EmailMaster",
}

# ── Header rows for auto-created tabs ────────────────────────────────────────

_SS_HEADERS = [
    "user_id", "month_year", "status", "completion_pct",
    "submitted_at", "locked_by", "locked_at", "checker_notes", "last_updated",
]

_RR_HEADERS = [
    "request_id", "zone_id", "location_id", "month_year",
    "reason", "status", "actioned_by", "actioned_at", "notes", "created_at",
]

_SETTINGS_HEADERS    = ["key", "value", "updated_by", "updated_at"]
_EMAIL_MASTER_HEADERS = ["type", "code", "name", "email", "cc"]

# Detail-table definitions: sheet_headers (Google Sheet columns) + data_keys (editable fields)
_DETAIL_DEF = {
    "RAILWAY_CLAIMS": {
        "sheet_headers": [
            "Sr#", "Zone", "Location", "Month-Year", "user_id",
            "Claim No.", "Year", "Amount (Rs)", "RR Nos.", "Ex", "To",
            "T/Wagon Nos.", "Product", "Qty.", "Rly.",
            "Pending Stage", "Status of Claim",
            "Last Hearing Date", "Next Hearing Date",
            "RCT Case Status as per Website", "Case Facts",
            "Rejection Reasons", "ShortComings/Discrepancies",
            "Strength of Case", "Recommendation",
        ],
        "data_keys": [
            "claim_no", "year", "amount", "rr_nos", "ex_station", "to_station",
            "wagon_nos", "product", "qty", "rly", "pending_stage", "status_claim",
            "last_hearing", "next_hearing", "rct_status", "case_facts",
            "rejection_reasons", "shortcomings", "strength", "recommendation",
        ],
        "prefix_count": 5,   # Sr#, Zone, Location, Month-Year, user_id
    },
    "IRR_DETAILS": {
        "sheet_headers": [
            "Sr#", "Zone", "Location Code", "Location Name", "Month-Year", "user_id",
            "IRR#", "IRR Date", "IRR Description", "IRR Amount (Rs)",
            "IRR Status (OPEN/CLOSED)", "IRR Closure Date",
        ],
        "data_keys": [
            "irr_no", "irr_date", "description", "amount", "status", "closure_date",
        ],
        "prefix_count": 6,   # Sr#, Zone, Location Code, Location Name, Month-Year, user_id
    },
    "LEGAL_CASES": {
        "sheet_headers": [
            "Sr. No", "Zone", "Location", "Month-Year", "user_id",
            "Court Name", "Case Number", "Cause Title", "Advocate",
            "Nature of Case", "Dealership Name and Location",
            "Background", "Status", "Last Hearing Date", "Next Hearing Date",
        ],
        "data_keys": [
            "court_name", "case_number", "cause_title", "advocate", "nature",
            "dealership", "background", "status", "last_hearing", "next_hearing",
        ],
        "prefix_count": 5,   # Sr. No, Zone, Location, Month-Year, user_id
    },
}

_MONTHS_S = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
MONTHS_LONG = ["January","February","March","April","May","June",
               "July","August","September","October","November","December"]


# ── Sheets client (cached for the app lifetime) ──────────────────────────────

@st.cache_resource
def _client():
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=_SCOPES
    )
    return gspread.authorize(creds)


@st.cache_resource
def _spreadsheet():
    return _client().open_by_key(st.secrets["sheets"]["spreadsheet_id"])


def mirror_backup_to_replica() -> dict:
    """Copy every tab of the live master sheet into the backup replica sheet.

    Full-value overwrite per tab (not incremental) -- safe to re-run any time.
    Takes roughly a minute (one read + one write per tab, ~25 tabs); intended
    to be triggered manually (Admin button) or by the scheduled GitHub Action,
    not on any hot path.
    """
    backup_id = st.secrets.get("sheets", {}).get("backup_spreadsheet_id", "")
    if not backup_id:
        return {"ok": False, "msg": "backup_spreadsheet_id not configured in secrets."}

    import time as _t
    t0 = _t.time()
    client  = _client()
    master  = _spreadsheet()
    try:
        replica = client.open_by_key(backup_id)
    except Exception as e:
        return {"ok": False, "msg": f"Could not open backup sheet: {e}"}

    existing = {ws.title: ws for ws in replica.worksheets()}
    tabs_done, errors = [], []
    master_tabs = master.worksheets()

    for i, ws in enumerate(master_tabs):
        name = ws.title
        try:
            rows = _api_call(ws.get_all_values)
            n_rows = max(len(rows), 1)
            n_cols = max((len(r) for r in rows), default=1) or 1

            if name in existing:
                dest = existing[name]
                if dest.row_count < n_rows or dest.col_count < n_cols:
                    _api_call(dest.resize, rows=max(n_rows, dest.row_count),
                             cols=max(n_cols, dest.col_count))
                _api_call(dest.clear)
            else:
                dest = _api_call(replica.add_worksheet, title=name,
                                 rows=max(n_rows, 10), cols=max(n_cols, 10))
                existing[name] = dest

            if rows:
                padded = [r + [""] * (n_cols - len(r)) for r in rows]
                _api_call(dest.update, padded, value_input_option="RAW")
            tabs_done.append(name)
        except Exception as e:
            errors.append(f"{name}: {e}")

        # Throttle: existing-tab writes use up to 3 write calls each (resize +
        # clear + update); back-to-back across ~25 tabs can exceed Sheets'
        # per-minute write quota even with _api_call's retry/backoff. A short
        # pause between tabs keeps total throughput comfortably under it.
        if i < len(master_tabs) - 1:
            _t.sleep(1.5)

    set_setting("last_backup_at", datetime.now().isoformat(), "system")
    return {
        "ok": not errors,
        "tabs_done": tabs_done,
        "errors": errors,
        "duration_sec": round(_t.time() - t0, 1),
    }


@st.cache_resource
def _ws_cache() -> dict:
    """App-lifetime dict {tab_name: Worksheet}.
    Avoids repeated ss.worksheet() API calls (each call re-fetches sheet list)."""
    return {}


def _ws(tab_name: str):
    cache = _ws_cache()
    if tab_name not in cache:
        cache[tab_name] = _spreadsheet().worksheet(tab_name)
    return cache[tab_name]


def _ensure_ws(tab_name: str, headers: list = None, force_headers: bool = False):
    """Return worksheet, auto-creating with headers if missing.
    When force_headers=True, also overwrites row 1 if it differs from headers."""
    cache = _ws_cache()
    if tab_name not in cache:
        ss = _spreadsheet()
        try:
            cache[tab_name] = _api_call(ss.worksheet, tab_name)
        except gspread.exceptions.WorksheetNotFound:
            cols = max(len(headers) if headers else 10, 26)
            ws   = _api_call(ss.add_worksheet, title=tab_name, rows=2000, cols=cols)
            if headers:
                _api_call(ws.append_row, headers, value_input_option="RAW")
            cache[tab_name] = ws
    ws = cache[tab_name]
    if force_headers and headers:
        existing = ws.get_all_values()
        if not existing or existing[0] != headers:
            ws.update("A1", [headers])
    return ws


def _field_label_map() -> dict:
    """Return {field_key: field_label} for all non-auto fields (from form_defs)."""
    try:
        from form_defs import SECTION_FIELDS
        return {
            f["key"]: f["label"]
            for fields in SECTION_FIELDS.values()
            for f in fields
            if not f.get("auto")
        }
    except Exception:
        return {}


# ── Month helpers ─────────────────────────────────────────────────────────────

def month_key(d: date = None) -> str:
    """date → 'Apr-2026'"""
    if d is None:
        d = date.today()
    return f"{_MONTHS_S[d.month - 1]}-{d.year}"


def parse_month_key(key: str):
    """'Apr-2026' → (month_0idx, year)"""
    parts = key.split("-")
    return _MONTHS_S.index(parts[0]), int(parts[1])


def compute_deadline(month_year: str) -> dict:
    """Return deadline date, days_left, urgency, and display strings for a month key."""
    month, year = parse_month_key(month_year)
    dl_month = month + 2          # +1 for 1-indexed, +1 for next month
    dl_year  = year
    if dl_month > 12:
        dl_month -= 12
        dl_year  += 1
    deadline  = date(dl_year, dl_month, 5)
    today     = date.today()
    days_left = (deadline - today).days

    if days_left < 0:
        urgency = "overdue"
    elif days_left <= 3:
        urgency = "urgent"
    elif days_left <= 7:
        urgency = "warning"
    else:
        urgency = "ok"

    return {
        "date":        f"{deadline.day}-{_MONTHS_S[deadline.month - 1]}-{deadline.year}",
        "days_left":   days_left,
        "urgency":     urgency,
        "month_label": MONTHS_LONG[month] + " " + str(year),
    }


def _prev_month_key() -> str:
    today = date.today()
    if today.month == 1:
        return month_key(date(today.year - 1, 12, 1))
    return month_key(date(today.year, today.month - 1, 1))


# ── Audit log (best-effort) ──────────────────────────────────────────────────

def audit_log(loc_code: str, action: str, details: str = ""):
    try:
        _pg_query(
            "insert into audit_log (actor_location_code, action, details) values (%s, %s, %s)",
            (loc_code, action, psycopg2.extras.Json(details)), fetch=False,
        )
    except Exception:
        pass


_IST_OFFSET = timedelta(hours=5, minutes=30)


def get_hourly_login_traffic(target_date: date) -> dict:
    """Return {"hours": [{"hour": 0..23, "users": n}, ...], "total_users": n}
    -- distinct users who logged in during each hour of target_date (IST), from
    audit_log's "Login" events, plus the distinct-user count for the whole day.

    Counts a user once per hour even if they logged in more than once in it --
    this answers "how many distinct users were on the portal that hour", not
    "how many login events fired". total_users is a separate day-wide distinct
    set (summing the per-hour counts would double-count anyone active in more
    than one hour).

    occurred_at is a real timestamptz (UTC) -- shift to IST before bucketing,
    same reasoning as the original's manual offset. Also now filters to the
    target date's UTC range in SQL instead of fetching the entire audit_log
    history and filtering in Python -- audit_log is the fastest-growing table
    in this system, so this matters more here than most of the other ports.
    """
    hour_users = {h: set() for h in range(24)}
    day_users: set = set()
    try:
        start_utc = datetime.combine(target_date, datetime.min.time(),
                                      tzinfo=timezone.utc) - _IST_OFFSET
        end_utc   = start_utc + timedelta(days=1)
        rows = _pg_query(
            "select occurred_at, actor_location_code from audit_log "
            "where action = 'Login' and occurred_at >= %s and occurred_at < %s",
            (start_utc, end_utc),
        )
        for r in rows:
            dt = r["occurred_at"]
            if dt is None:
                continue
            dt_ist = dt.astimezone(timezone.utc) + _IST_OFFSET
            loc = r["actor_location_code"] or ""
            hour_users[dt_ist.hour].add(loc)
            day_users.add(loc)
    except Exception:
        pass
    return {
        "hours": [{"hour": h, "users": len(hour_users[h])} for h in range(24)],
        "total_users": len(day_users),
    }


# ── Location name lookup from master ─────────────────────────────────────────

@st.cache_data(ttl=3600)
def _loc_name_map() -> dict:
    """Return {location_code_upper: (location_name, loc_type, zone)} from Postgres
    locations/zones. Same shape/contract as the original Sheets-backed version --
    loc_type comes straight from the locations.loc_type check constraint now
    instead of being pattern-guessed from the name."""
    try:
        rows = _pg_query("""
            select l.code, l.name, l.loc_type, z.name as zone_name
            from locations l
            left join zones z on z.id = l.zone_id
            where l.active
        """)
        return {
            r["code"].upper(): (r["name"], r["loc_type"], r["zone_name"] or "")
            for r in rows
        }
    except Exception:
        return {}


def get_loc_type(loc_code: str) -> str:
    """Return loc_type ('HPCL'|'TOP'|'HMEL') for a location code."""
    _, ltype = _resolve_loc_info(loc_code, loc_code)
    return ltype


def reset_location_data(loc_code: str) -> dict:
    """Delete ALL MIS data for a location (pre-launch data cleanup).

    Clears: MIS_DRAFT, SubmissionStatus, Railway_Claims, IRR_Details,
    Legal_Cases, MI_TANK_OUTAGE, MI_MAJOR_REPAIR, MI_VRU, MI_AUDIT_2526,
    MI_AUDIT_2627, MI_TECH_AUDIT, MI_EQUIP_BREAKDOWN, MI_INT_PIPELINE,
    MI_EXT_PIPELINE, MI_TANK_STATUS.
    """
    uid = str(loc_code).strip()
    tabs_to_clear = [
        "MIS_DRAFT", "SUBMISSION_STATUS",
        "RAILWAY_CLAIMS", "IRR_DETAILS", "LEGAL_CASES",
        "MI_TANK_OUTAGE", "MI_MAJOR_REPAIR", "MI_VRU",
        "MI_AUDIT_2526", "MI_AUDIT_2627", "MI_TECH_AUDIT",
        "MI_EQUIP_BREAKDOWN", "MI_INT_PIPELINE", "MI_EXT_PIPELINE",
        "MI_TANK_STATUS",
    ]
    deleted_total = 0
    errors = []
    try:
        for tab_key in tabs_to_clear:
            tab_name = TABS.get(tab_key)
            if not tab_name:
                continue
            try:
                ws = _ws(tab_name)
                rows_to_del = []
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if row and str(row[0]).strip() == uid:
                        rows_to_del.append(i)
                for row_idx in reversed(rows_to_del):
                    ws.delete_rows(row_idx)
                deleted_total += len(rows_to_del)
            except Exception as tab_err:
                errors.append(f"{tab_name}: {tab_err}")
        # Also clear the cached dashboard data
        try:
            get_dashboard_data.clear()
        except Exception:
            pass
        msg = f"Deleted {deleted_total} row(s) for location {uid}."
        if errors:
            msg += "  Warnings: " + "; ".join(errors)
        return {"ok": True, "deleted": deleted_total, "msg": msg}
    except Exception as exc:
        return {"ok": False, "msg": str(exc)}


def _resolve_loc_info(loc_code: str, stored_name: str) -> tuple:
    """Return (loc_name, loc_type) for a location code."""
    code_up = loc_code.strip().upper()
    master  = _loc_name_map()
    if code_up in master:
        entry = master[code_up]
        name  = entry[0]
        ltype = entry[1]
        final_name = stored_name.strip() if (
            stored_name.strip() and stored_name.strip().upper() != code_up
        ) else name
        return final_name, ltype
    # Not in master: derive loc_type from name/code
    nu, cu = stored_name.upper(), code_up
    if "HMEL" in nu or "HMEL" in cu:
        ltype = "HMEL"
    elif "TOP" in nu or "JAMNAGAR" in nu or "TOP" in cu:
        ltype = "TOP"
    else:
        ltype = "HPCL"
    return stored_name.strip() or loc_code.strip(), ltype


def _resolve_loc_name(loc_code: str, stored_name: str) -> str:
    """Backward-compat wrapper — returns loc_name only."""
    return _resolve_loc_info(loc_code, stored_name)[0]


# ── Authentication ────────────────────────────────────────────────────────────

def check_login(location_code: str, password: str) -> dict:
    """Postgres version. Same return shape as the original except the two
    fields (_sheet_row, _password) confirmed unused anywhere in app.py are
    dropped. Passwords are bcrypt-hashed in `users.password_hash` (migrated
    from the sheet's plaintext by import_user_access.mjs) -- verified with
    bcrypt.checkpw instead of a plain string comparison."""
    try:
        location_code = str(location_code or "").strip()
        password      = str(password      or "").strip()

        if not location_code or not password:
            return {"ok": False, "msg": "Location Code and Password are required."}

        row = _pg_one("""
            select u.id, u.login_code, u.password_hash, u.role, u.is_first_login,
                   u.location_code, u.active,
                   l.name as loc_name, l.loc_type,
                   z.name as zone_name
            from users u
            left join locations l on l.code = u.location_code
            left join zones z on z.id = u.zone_id
            where upper(u.login_code) = upper(%s)
        """, (location_code,))

        if not row or not row["active"]:
            return {"ok": False, "msg": "Location Code not found. Please check and try again."}

        if not bcrypt.checkpw(password.encode(), row["password_hash"].encode()):
            return {"ok": False, "msg": "Incorrect password. Please try again."}

        _pg_query("update users set last_login_at = now() where id = %s",
                   (row["id"],), fetch=False)

        role = row["role"] or "Maker"
        audit_log(row["login_code"], "Login", f"Successful login as {role}")

        if role in ("Maker", "Checker"):
            loc_name, loc_type = row["loc_name"] or row["login_code"], row["loc_type"] or "HPCL"
        elif role == "Zone":
            loc_name, loc_type = row["zone_name"] or row["login_code"], "HPCL"
        else:
            loc_name, loc_type = row["login_code"], "HPCL"

        return {
            "ok":           True,
            "userId":       row["login_code"],
            "locName":      loc_name,
            "locType":      loc_type,   # HPCL | TOP | HMEL
            "zone":         row["zone_name"] or "",
            "role":         role,
            "isFirstLogin": bool(row["is_first_login"]),
        }

    except Exception as e:
        # Don't leak internal details (connection strings, stack text) to the
        # login screen -- log server-side for diagnosis, show a plain message.
        print(f"[check_login] {type(e).__name__}: {e}")
        return {"ok": False,
                "msg": "Unable to reach the server right now. Please check your "
                       "connection and try again in a moment."}


def change_password(user_id: str, current_pass: str,
                    new_pass: str, confirm_pass: str) -> dict:
    try:
        user_id      = str(user_id      or "").strip()
        current_pass = str(current_pass or "").strip()
        new_pass     = str(new_pass     or "").strip()
        confirm_pass = str(confirm_pass or "").strip()

        if len(new_pass) < 6:
            return {"ok": False, "msg": "New password must be at least 6 characters."}
        if new_pass != confirm_pass:
            return {"ok": False, "msg": "New password and Confirm password do not match."}
        if new_pass == current_pass:
            return {"ok": False, "msg": "New password must be different from the current password."}

        row = _pg_one(
            "select id, password_hash, role, location_code from users where login_code = %s",
            (user_id,),
        )
        if not row:
            return {"ok": False, "msg": "User record not found."}
        if not bcrypt.checkpw(current_pass.encode(), row["password_hash"].encode()):
            return {"ok": False, "msg": "Current password is incorrect."}

        new_hash = bcrypt.hashpw(new_pass.encode(), bcrypt.gensalt()).decode()
        _pg_query(
            """update users set password_hash = %s, is_first_login = false,
               last_password_change_at = now() where id = %s""",
            (new_hash, row["id"]), fetch=False,
        )
        audit_log(user_id, "Password Changed",
                  f"Password changed successfully. IsFirstLogin reset to FALSE. "
                  f"Role: {row['role'] or 'Maker'}. Location: {row['location_code'] or ''}")
        return {"ok": True}

    except Exception as e:
        return {"ok": False, "msg": f"System error: {e}"}


# ── Help desk / forgot password ──────────────────────────────────────────────

def request_password_reset(location_code: str) -> dict:
    """Postgres note: helpdesk_tickets.location_code has a not-null FK to
    locations -- Zone/Admin/Viewer users have no location_code at all, so
    they can't get a helpdesk ticket row the way Maker/Checker can. Falls
    back to audit_log only for those; a narrow edge case in practice, since
    self-service reset is used almost entirely by Maker/Checker."""
    try:
        location_code = str(location_code or "").strip()
        if not location_code:
            return {"ok": False, "msg": "Please enter your Location Code."}

        user = _pg_one("select location_code from users where login_code = %s", (location_code,))
        if not user:
            return {"ok": False, "msg": f'Location Code "{location_code}" is not registered.'}

        if user["location_code"]:
            _pg_query(
                "insert into helpdesk_tickets (location_code, issue_type, issue_desc, status) "
                "values (%s, 'Password Reset Request', "
                "'User requested a password reset via the Forgot Password link.', 'OPEN')",
                (user["location_code"],), fetch=False,
            )
        audit_log(location_code, "Forgot Password", "Reset request logged")
        return {
            "ok":  True,
            "msg": "Your password reset request has been logged. "
                   "Admin will reset your password shortly.",
        }
    except Exception as e:
        return {"ok": False, "msg": f"System error: {e}"}


_HELPDESK_HEADERS = [
    "timestamp", "location_code", "issue_type",
    "issue_desc", "status", "admin_response", "responded_at",
]


# Sheets used free-text status ("Pending"/"In Progress"/"Resolved", see
# app.py's selectbox); schema constrains to OPEN/RESPONDED/CLOSED. Map
# between them rather than loosen the constraint.
_HD_STATUS_TO_DB   = {"Pending": "OPEN", "In Progress": "RESPONDED", "Resolved": "CLOSED"}
_HD_STATUS_FROM_DB = {v: k for k, v in _HD_STATUS_TO_DB.items()}


def get_helpdesk_tickets() -> list:
    """Return all helpdesk tickets (newest first).

    'row' is now the ticket's real id, used by respond_to_helpdesk_ticket
    (no longer a literal sheet row number, but the same round-trip role)."""
    try:
        rows = _pg_query("""
            select id, created_at, location_code, issue_type, issue_desc,
                   status, admin_response, responded_at
            from helpdesk_tickets order by created_at desc
        """)
        return [
            {
                "row":            r["id"],
                "timestamp":      r["created_at"].isoformat() if r["created_at"] else "",
                "location_code":  r["location_code"] or "",
                "issue_type":     r["issue_type"] or "",
                "issue_desc":     r["issue_desc"] or "",
                "status":         _HD_STATUS_FROM_DB.get(r["status"], "Pending"),
                "admin_response": r["admin_response"] or "",
                "responded_at":   r["responded_at"].isoformat() if r["responded_at"] else "",
            }
            for r in rows
        ]
    except Exception:
        return []


def respond_to_helpdesk_ticket(row: int, response: str,
                                status: str, updated_by: str) -> dict:
    """Write admin response + status back to the ticket (row = ticket id)."""
    try:
        db_status = _HD_STATUS_TO_DB.get(status, "RESPONDED")
        actioner  = _pg_one("select id from users where login_code = %s", (updated_by,))
        _pg_query(
            "update helpdesk_tickets set status = %s, admin_response = %s, "
            "responded_at = now(), responded_by = %s where id = %s",
            (db_status, response, actioner["id"] if actioner else None, row), fetch=False,
        )
        audit_log(updated_by, "HelpDesk Response",
                  f"Row {row} → {status}: {response[:80]}")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def log_help_request(location_code: str, issue_desc: str,
                     issue_type: str = "Help Request") -> dict:
    """FLAGGED FOR REVIEW (minor): helpdesk_tickets.location_code has a real
    foreign key to locations(code), so a mistyped/unknown code is now
    rejected with a friendly message instead of being silently logged as-is
    (the Sheets version never validated this at all). Reasonable tightening,
    not a regression, but worth knowing it's a new behavior."""
    try:
        location_code = str(location_code or "").strip()
        issue_desc    = str(issue_desc    or "").strip()
        issue_type    = str(issue_type    or "Help Request").strip()

        if not location_code:
            return {"ok": False, "msg": "Please enter your Location Code."}
        if len(issue_desc) < 10:
            return {"ok": False, "msg": "Please describe your issue in at least 10 characters."}
        if not _pg_one("select 1 from locations where code = %s", (location_code,)):
            return {"ok": False, "msg": f'Location Code "{location_code}" is not recognized.'}

        ref_id = "HD-" + str(int(time.time()))[-6:]
        _pg_query(
            "insert into helpdesk_tickets (location_code, issue_type, issue_desc, status) "
            "values (%s, %s, %s, 'OPEN')",
            (location_code, issue_type, issue_desc), fetch=False,
        )
        audit_log(location_code, "Help Request", f"{ref_id} [{issue_type}] — {issue_desc[:60]}")
        return {
            "ok":  True,
            "ref": ref_id,
            "msg": f"Ticket raised (Ref: {ref_id}). Admin will respond shortly.",
        }
    except Exception as e:
        return {"ok": False, "msg": f"System error: {e}"}


# ── Submission status ─────────────────────────────────────────────────────────

def _revert_if_deleted(user_id: str, month_year: str, status: str,
                       completion_pct: float) -> str:
    """If the approved_snapshots row was deleted externally, revert status
    → IN_PROGRESS. Postgres equivalent of the Sheets version's "did the
    MIS_Submitted row disappear" check -- approved_snapshots is the table
    that gets a row on approval, same role MIS_Submitted played."""
    if status not in ("SUBMITTED", "LOCKED"):
        return status
    sub = _pg_one(
        "select id from monthly_submissions where location_code = %s and month_year = %s",
        (user_id, _mk_to_date(month_year)),
    )
    exists = bool(sub) and bool(_pg_one(
        "select 1 from approved_snapshots where submission_id = %s", (sub["id"],)
    ))
    if not exists:
        _update_submission_status(user_id, month_year, "IN_PROGRESS", completion_pct)
        return "IN_PROGRESS"
    return status


def get_month_status(user_id: str, month_year: str) -> dict:
    """No caching here (unlike the Sheets version's 60s @st.cache_data) --
    this is exactly the class of read where a stale cache previously caused
    real confusion (status appearing to "revert"). Postgres reads are cheap
    enough that correctness wins over shaving a query."""
    try:
        row = _pg_one(
            "select status, completion_pct, checker_notes from monthly_submissions "
            "where location_code = %s and month_year = %s",
            (user_id, _mk_to_date(month_year)),
        )
        if row:
            pct    = float(row["completion_pct"] or 0)
            status = _revert_if_deleted(user_id, month_year, row["status"] or "NOT_STARTED", pct)
            return {
                "status":         status,
                "completion_pct": pct,
                "is_locked":      status in ("SUBMITTED", "LOCKED", "PENDING_REVIEW"),
                "checker_notes":  row["checker_notes"] or "",
            }
    except Exception:
        pass
    return {"status": "NOT_STARTED", "completion_pct": 0.0, "is_locked": False}


def get_fy_months(user_id: str, fy_start_year: int) -> dict:
    """Return all 12 months of a financial year (Apr→Mar) with submission status."""
    try:
        user_id = str(user_id or "").strip()

        status_map: dict = {}
        pct_map:    dict = {}
        try:
            rows = _pg_query(
                "select month_year, status, completion_pct from monthly_submissions "
                "where location_code = %s",
                (user_id,),
            )
            for r in rows:
                k = month_key(r["month_year"])
                status_map[k] = r["status"]
                pct_map[k]    = float(r["completion_pct"] or 0)
        except Exception:
            pass

        months = []
        for i in range(12):
            # i=0→April, i=1→May, …, i=8→December, i=9→January, …, i=11→March
            month_num = (i + 3) % 12 + 1
            yr        = fy_start_year if i < 9 else fy_start_year + 1
            d         = date(yr, month_num, 1)
            key       = month_key(d)
            label     = MONTHS_LONG[d.month - 1] + " " + str(d.year)
            status    = _revert_if_deleted(
                user_id, key, status_map.get(key, "NOT_STARTED"), pct_map.get(key, 0.0)
            )
            months.append({
                "value":     key,
                "label":     label,
                "status":    status,
                "is_locked": status in ("SUBMITTED", "LOCKED", "PENDING_REVIEW"),
            })

        return {"ok": True, "months": months}

    except Exception as e:
        return {"ok": False, "msg": str(e), "months": []}


def get_available_months(user_id: str) -> dict:
    try:
        user_id = str(user_id or "").strip()
        today   = date.today()

        # Fetch submission status — empty map on any error
        status_map: dict = {}
        try:
            rows = _pg_query(
                "select month_year, status from monthly_submissions where location_code = %s",
                (user_id,),
            )
            status_map = {month_key(r["month_year"]): r["status"] for r in rows}
        except Exception:
            pass

        months = []
        for m in range(1, 13):
            month_0 = today.month - 1 - m   # 0-indexed month
            yr      = today.year
            while month_0 < 0:
                month_0 += 12
                yr      -= 1
            d      = date(yr, month_0 + 1, 1)
            key    = month_key(d)
            label  = MONTHS_LONG[d.month - 1] + " " + str(d.year)
            status = status_map.get(key, "NOT_STARTED")
            months.append({
                "value":     key,
                "label":     label,
                "status":    status,
                "is_locked": status in ("SUBMITTED", "LOCKED", "PENDING_REVIEW"),
            })

        return {"ok": True, "months": months}

    except Exception as e:
        return {"ok": False, "msg": str(e), "months": []}


@st.cache_data(ttl=60, show_spinner=False)
def get_dashboard_data(user_id: str, month_year: str = None, loc_type: str = "HPCL") -> dict:
    try:
        user_id = str(user_id or "").strip()
        if not user_id:
            return {"ok": False, "msg": "Session expired. Please log in again."}

        if not month_year:
            month_year = _prev_month_key()

        status_data = get_month_status(user_id, month_year)
        deadline    = compute_deadline(month_year)

        # Derive per-section completion from the draft row (column C)
        secs_done: list = []
        draft: dict = {}
        try:
            draft    = load_draft(user_id, month_year)
            secs_raw = draft.get("_sections_complete", "")
            secs_done = sorted(
                {int(x) for x in secs_raw.split(",") if x.strip().isdigit()}
            )
        except Exception:
            pass

        # Validate stored completion against current SECTION_FIELDS definitions.
        # A section is only complete if every current required non-auto non-excluded
        # field has a non-empty value in the saved draft.
        try:
            from form_defs import SECTION_FIELDS, get_excluded_fields, get_skip_sections
            excl_keys  = get_excluded_fields(loc_type)
            skip_secs  = get_skip_sections(loc_type)
            valid_secs = []
            for sec_num in secs_done:
                if sec_num in skip_secs:
                    valid_secs.append(sec_num)  # skip-sections always count as done
                    continue
                all_filled = True
                for f in SECTION_FIELDS.get(sec_num, []):
                    if f.get("auto") or not f.get("req"):
                        continue
                    if f["key"] in excl_keys:
                        continue
                    # Conditional field — skip if its show_when condition is not met
                    sw = f.get("show_when")
                    if sw and not all(
                        str(draft.get(k) or "") == str(v) for k, v in sw.items()
                    ):
                        continue
                    if draft.get(f["key"]) in (None, ""):
                        all_filled = False
                        break
                if all_filled:
                    valid_secs.append(sec_num)
            secs_done = sorted(valid_secs)
        except Exception:
            pass  # on any import/logic error keep original secs_done

        # S5 only counts as complete when M&I MIS (S5A) is also fully filled
        mi_complete = check_mi_complete(user_id, month_year) if 5 in secs_done else False
        if 5 in secs_done and not mi_complete:
            secs_done = [s for s in secs_done if s != 5]

        eff_pct = len(secs_done) * 10.0

        return {
            "ok":             True,
            "month_year":     month_year,
            "status":         status_data["status"],
            "completion_pct": eff_pct,
            "is_locked":      status_data["is_locked"],
            "checker_notes":  status_data.get("checker_notes", ""),
            "secs_done":      secs_done,
            "mi_complete":    mi_complete,
            "deadline":       deadline,
        }
    except Exception as e:
        return {"ok": False, "msg": f"Error loading dashboard: {e}"}


# ── Draft CRUD ────────────────────────────────────────────────────────────────
# Postgres: field_values(submission_id, field_key, value), one row per field.
# sections_complete is stored as field_key='_sections_complete' in the same
# table -- the original already treated it as "just another key" in the
# in-memory dict (see load_draft below); this keeps that, without a schema
# change for one extra text column.

def load_draft(user_id: str, month_year: str) -> dict:
    """Return draft data dict for user+month, or empty dict."""
    try:
        sub = _pg_one(
            "select id from monthly_submissions where location_code = %s and month_year = %s",
            (user_id, _mk_to_date(month_year)),
        )
        if not sub:
            return {}
        rows = _pg_query(
            "select field_key, value from field_values where submission_id = %s",
            (sub["id"],),
        )
        data = {r["field_key"]: r["value"] for r in rows}
        data.setdefault("_sections_complete", "")
        data["_sheet_row"] = sub["id"]   # kept for compatibility; not a real row number anymore
        return data
    except Exception:
        return {}


def load_submitted_fields(user_id: str, month_year: str) -> dict:
    """Read field key→value dict from the approved snapshot for a SUBMITTED
    location. Fallback for generate_filled_mis_report when the draft is
    unavailable or empty. Postgres note: approved_snapshots.snapshot is
    already a field_key→value jsonb blob (written at approval time) -- no
    label reverse-mapping needed, unlike the Sheets version which had to
    map MIS_Submitted's human-readable column labels back to f1/f2/... keys."""
    try:
        row = _pg_one("""
            select aps.snapshot
            from approved_snapshots aps
            join monthly_submissions ms on ms.id = aps.submission_id
            where ms.location_code = %s and ms.month_year = %s
        """, (user_id, _mk_to_date(month_year)))
        return dict(row["snapshot"]) if row and row["snapshot"] else {}
    except Exception:
        return {}


def _update_submission_status(user_id: str, month_year: str, status: str, pct: float,
                               submitted_at: str = "", locked_by: str = "",
                               locked_at: str = "", checker_notes: str = ""):
    """THE fix this whole migration started over. The Sheets version reads
    the (15s-cached) sheet, decides update-vs-append in Python, and writes --
    two calls landing in that window both see "no row yet" and both append,
    producing the duplicate SubmissionStatus rows the audit found (9 real
    cases, e.g. 1708/Apr-2026 and 1708/Jun-2026 stuck showing IN_PROGRESS
    despite being actually SUBMITTED).

    This is a single INSERT ... ON CONFLICT ... DO UPDATE -- atomic at the
    database level via the unique(location_code, month_year) constraint.
    Two concurrent callers can't both "win" an append; Postgres serializes
    them and the second becomes an UPDATE. The race is structurally
    impossible, not just less likely.
    """
    mdate = _mk_to_date(month_year)
    locked_by_id = None
    if locked_by:
        row = _pg_one("select id from users where login_code = %s", (locked_by,))
        locked_by_id = row["id"] if row else None

    _pg_query(
        """
        insert into monthly_submissions
            (location_code, month_year, status, completion_pct, submitted_at,
             locked_by, locked_at, checker_notes, last_updated_at)
        values (%(uid)s, %(mdate)s, %(status)s, %(pct)s,
                coalesce(%(sub_at)s::timestamptz, now()),
                %(locked_by_id)s, %(lock_at)s::timestamptz, %(notes)s, now())
        on conflict (location_code, month_year) do update set
            status          = excluded.status,
            completion_pct  = excluded.completion_pct,
            submitted_at    = coalesce(%(sub_at)s::timestamptz, monthly_submissions.submitted_at),
            locked_by       = coalesce(%(locked_by_id)s, monthly_submissions.locked_by),
            locked_at       = coalesce(%(lock_at)s::timestamptz, monthly_submissions.locked_at),
            checker_notes   = coalesce(nullif(%(notes)s, ''), monthly_submissions.checker_notes),
            last_updated_at = now()
        """,
        {
            "uid": user_id, "mdate": mdate, "status": status, "pct": pct,
            "sub_at": submitted_at or None, "locked_by_id": locked_by_id,
            "lock_at": locked_at or None, "notes": checker_notes,
        },
        fetch=False,
    )


def save_draft(user_id: str, month_year: str,
               section_num: int | None = None,
               field_data: dict | None = None,
               mark_complete: bool = False,
               sections_complete: list | None = None) -> dict:
    """Merge field_data into existing draft and persist.

    Two calling modes:
      Normal (per-section save):
        save_draft(user_id, month_year, section_num, field_data, mark_complete)
      Bulk upload save:
        save_draft(user_id, month_year, field_data={...}, sections_complete=[1,3,5,...])

    Postgres note: the Sheets version's "fresh, never trust the cache"
    comment described working around a duplicate-row race on MIS_DRAFT.
    Here, the monthly_submissions upsert is a single atomic statement
    (unique constraint + ON CONFLICT, same pattern as
    _update_submission_status) and field_values writes are keyed by
    (submission_id, field_key) primary key -- no equivalent race exists to
    guard against.
    """
    try:
        existing = load_draft(user_id, month_year)
        secs_raw = existing.get("_sections_complete", "")
        try:
            secs_done = {int(x) for x in secs_raw.split(",") if x.strip().isdigit()}
        except Exception:
            secs_done = set()

        merged = {k: v for k, v in existing.items() if not k.startswith("_")}
        if field_data:
            merged.update(field_data)

        if sections_complete is not None:
            # Bulk upload: replace section completion entirely
            secs_done = set(sections_complete)
        elif section_num is not None:
            if mark_complete:
                secs_done.add(section_num)
            else:
                secs_done.discard(section_num)

        pct      = len(secs_done) * 10.0
        secs_str = ",".join(str(s) for s in sorted(secs_done))
        mdate    = _mk_to_date(month_year)

        sub = _pg_one(
            """
            insert into monthly_submissions (location_code, month_year)
            values (%s, %s)
            on conflict (location_code, month_year) do update set
                last_updated_at = now()
            returning id, status
            """,
            (user_id, mdate),
        )
        submission_id = sub["id"]

        merged["_sections_complete"] = secs_str
        for key, val in merged.items():
            _pg_query(
                """
                insert into field_values (submission_id, field_key, value)
                values (%s, %s, %s)
                on conflict (submission_id, field_key) do update set
                    value = excluded.value, updated_at = now()
                """,
                (submission_id, key, str(val) if val is not None else ""),
                fetch=False,
            )

        # Never let an incidental data save silently un-approve a month the
        # Checker has already reviewed (same guard as the Sheets version --
        # this specific check still matters regardless of backend).
        if sub["status"] not in ("SUBMITTED", "LOCKED"):
            status = "IN_PROGRESS" if secs_done else "NOT_STARTED"
            _update_submission_status(user_id, month_year, status, pct)

        tag = f"S{section_num}" if section_num else "BulkUpload"
        audit_log(user_id, f"SaveDraft {tag}",
                  f"month={month_year} secs={secs_str} pct={pct}")

        return {"ok": True, "pct": pct, "secs_done": sorted(secs_done)}

    except Exception as e:
        return {"ok": False, "msg": str(e)}


# ── Detail table CRUD (Railway Claims, IRR Details, Legal Cases) ─────────────

# tab_key (TABS dict convention) -> detail_rows.table_type (schema check constraint)
_DETAIL_TYPE_MAP = {
    "RAILWAY_CLAIMS": "RAILWAY_CLAIM",
    "IRR_DETAILS":    "IRR_DETAIL",
    "LEGAL_CASES":    "LEGAL_CASE",
}


def load_detail_table(user_id: str, month_year: str, tab_key: str) -> list:
    """Return list of row-dicts for user+month from a detail table.

    Postgres note: Sr#/Zone/Location prefix columns the sheet stored per row
    are gone -- those were denormalized display convenience, derivable from
    the submission's own location/month via a join if a report ever needs
    them. row_data holds exactly defn["data_keys"], nothing else."""
    try:
        defn  = _DETAIL_DEF[tab_key]
        ttype = _DETAIL_TYPE_MAP[tab_key]
        sub = _pg_one(
            "select id from monthly_submissions where location_code = %s and month_year = %s",
            (user_id, _mk_to_date(month_year)),
        )
        if not sub:
            return []
        rows = _pg_query(
            "select row_data from detail_rows where submission_id = %s and table_type = %s "
            "order by sort_order",
            (sub["id"], ttype),
        )
        return [{k: r["row_data"].get(k, "") for k in defn["data_keys"]} for r in rows]
    except Exception:
        return []


def save_detail_table(user_id: str, month_year: str, tab_key: str,
                      rows_data: list, user_info: dict) -> dict:
    """Replace all rows for user+month in a detail table with rows_data."""
    try:
        defn  = _DETAIL_DEF[tab_key]
        ttype = _DETAIL_TYPE_MAP[tab_key]
        sub = _pg_one(
            """
            insert into monthly_submissions (location_code, month_year)
            values (%s, %s)
            on conflict (location_code, month_year) do update set last_updated_at = now()
            returning id
            """,
            (user_id, _mk_to_date(month_year)),
        )
        submission_id = sub["id"]

        _pg_query(
            "delete from detail_rows where submission_id = %s and table_type = %s",
            (submission_id, ttype), fetch=False,
        )
        for sr, rec in enumerate(rows_data, 1):
            row_data = {k: str(rec.get(k, "") or "") for k in defn["data_keys"]}
            _pg_query(
                "insert into detail_rows (submission_id, table_type, row_data, sort_order) "
                "values (%s, %s, %s, %s)",
                (submission_id, ttype, psycopg2.extras.Json(row_data), sr), fetch=False,
            )

        audit_log(user_id, f"SaveDetail {tab_key}",
                  f"month={month_year} rows={len(rows_data)}")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


# ── Maker-Checker Submit / Approve / Reject workflow ─────────────────────────

def submit_for_review(user_id: str, month_year: str) -> dict:
    """Maker submits completed draft for Checker review."""
    try:
        from form_defs import get_skip_sections
        loc_type      = get_loc_type(user_id)
        skip_secs     = get_skip_sections(loc_type)
        required_secs = {n for n in range(1, 11) if n not in skip_secs}

        sd        = get_month_status(user_id, month_year)
        pct       = sd.get("completion_pct", 0)
        dash      = get_dashboard_data(user_id, month_year, loc_type)
        secs_done = set(dash.get("secs_done", []))

        if not required_secs.issubset(secs_done):
            missing = sorted(required_secs - secs_done)
            return {"ok": False,
                    "msg": ("Cannot submit — sections "
                            f"{', '.join('S' + str(n) for n in missing)} "
                            "must be saved first.")}
        if 5 not in skip_secs and not check_mi_complete(user_id, month_year):
            return {"ok": False,
                    "msg": ("Cannot submit — M&I MIS (S5A) is incomplete. "
                            "Please fill all 10 tabs in S5A and save each one.")}
        if sd["status"] in ("SUBMITTED", "LOCKED", "PENDING_REVIEW"):
            return {"ok": False, "msg": f"Month is already {sd['status'].replace('_',' ').lower()}."}
        _update_submission_status(user_id, month_year, "PENDING_REVIEW", pct,
                                  submitted_at=datetime.now().isoformat())
        audit_log(user_id, "SubmitForReview", f"month={month_year}")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def approve_submission(maker_id: str, month_year: str, checker_id: str,
                       flat_data: dict, user_info: dict) -> dict:
    """Checker approves — writes an approved snapshot and locks the month.

    Postgres note: approved_snapshots.snapshot stores field_key->value
    directly (jsonb) -- no label mapping needed at write time the way the
    Sheets version needed to convert field keys to human-readable column
    headers for MIS_Submitted. Report generation (Domain 11) does that
    mapping at read/render time instead, a cleaner separation."""
    try:
        mdate = _mk_to_date(month_year)
        sub = _pg_one(
            "select id from monthly_submissions where location_code = %s and month_year = %s",
            (maker_id, mdate),
        )
        if not sub:
            return {"ok": False, "msg": "No submission found to approve."}

        snapshot = {k: (str(v) if v is not None else "") for k, v in flat_data.items()}
        _pg_query(
            """
            insert into approved_snapshots
                (submission_id, location_code, month_year, snapshot, approved_by, approved_at)
            values (%s, %s, %s, %s, (select id from users where login_code = %s), now())
            on conflict (submission_id) do update set
                snapshot = excluded.snapshot, approved_by = excluded.approved_by, approved_at = now()
            """,
            (sub["id"], maker_id, mdate, psycopg2.extras.Json(snapshot), checker_id),
            fetch=False,
        )

        sd = get_month_status(maker_id, month_year)
        _update_submission_status(maker_id, month_year, "SUBMITTED",
                                  sd.get("completion_pct", 100),
                                  locked_by=checker_id, locked_at=datetime.now().isoformat())
        audit_log(checker_id, "ApproveSubmission", f"maker={maker_id} month={month_year}")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def reject_submission(maker_id: str, month_year: str,
                      checker_id: str, note: str) -> dict:
    """Checker rejects — returns to editable state with a note for the Maker."""
    try:
        sd = get_month_status(maker_id, month_year)
        if sd["status"] != "PENDING_REVIEW":
            return {"ok": False, "msg": "Only PENDING_REVIEW submissions can be rejected."}
        _update_submission_status(maker_id, month_year, "REJECTED",
                                  sd.get("completion_pct", 0),
                                  checker_notes=note)
        audit_log(checker_id, "RejectSubmission",
                  f"maker={maker_id} month={month_year} note={note[:60]}")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def reset_draft(maker_id: str, month_year: str, checker_id: str, reason: str) -> dict:
    """Checker resets a maker's draft — wipes all field data so maker can start fresh."""
    try:
        sub = _pg_one(
            "select id from monthly_submissions where location_code = %s and month_year = %s",
            (maker_id, _mk_to_date(month_year)),
        )
        if sub:
            _pg_query("delete from field_values where submission_id = %s",
                      (sub["id"],), fetch=False)

        _update_submission_status(
            maker_id, month_year, "NOT_STARTED", 0.0,
            checker_notes=f"[RESET by {checker_id}] {reason}",
        )
        audit_log(checker_id, "ResetDraft",
                  f"maker={maker_id} month={month_year} reason={reason[:80]}")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


# ── Phase-6: Zone / HQO view + Revision workflow ──────────────────────────────

@st.cache_data(ttl=60, show_spinner=False)
def _user_access_raw_rows() -> list:
    """Shared 60 s cache of the full UserAccess sheet -- avoids an uncached
    full-sheet read on every rerun of every Zone/Admin dashboard render,
    which was a real contributor to the app feeling slow with 15-20+
    concurrent users (Streamlit reruns the whole script on every widget
    interaction, and this sheet was being re-read from scratch each time)."""
    ws = _ws(TABS["USER_ACCESS"])
    return _api_call(ws.get_all_values)


def get_locations_by_zone(zone_name: str) -> list:
    """Return Maker location dicts for a given zone."""
    try:
        rows = _pg_query("""
            select u.login_code, l.name as loc_name, z.name as zone_name
            from users u
            join locations l on l.code = u.location_code
            join zones z on z.id = l.zone_id
            where u.role = 'Maker' and z.name = %s
        """, (zone_name,))
        return [{"userId": r["login_code"], "locName": r["loc_name"], "zone": r["zone_name"]}
                for r in rows]
    except Exception:
        return []


def get_all_maker_locations() -> list:
    """Return all Maker location dicts across all zones (HQO view)."""
    try:
        rows = _pg_query("""
            select u.login_code, l.name as loc_name, z.name as zone_name
            from users u
            join locations l on l.code = u.location_code
            left join zones z on z.id = l.zone_id
            where u.role = 'Maker'
        """)
        return [{"userId": r["login_code"], "locName": r["loc_name"], "zone": r["zone_name"] or ""}
                for r in rows]
    except Exception:
        return []


def get_maker_info(user_id: str) -> dict:
    """Return {"userId", "locName", "zone"} for a Maker, or a minimal dict."""
    try:
        row = _pg_one("""
            select u.login_code, l.name as loc_name, z.name as zone_name
            from users u
            join locations l on l.code = u.location_code
            left join zones z on z.id = l.zone_id
            where u.role = 'Maker' and u.login_code = %s
        """, (user_id,))
        if row:
            return {"userId": row["login_code"], "locName": row["loc_name"],
                    "zone": row["zone_name"] or ""}
    except Exception:
        pass
    return {"userId": user_id, "locName": user_id, "zone": ""}


def get_submissions_for_locations(locs: list, month_year: str) -> list:
    """Bulk-fetch submission status for a list of location dicts.

    Locations in the 'excluded_from_reports' Settings key (e.g. non-operational)
    are silently omitted from the returned list.
    """
    excluded = get_excluded_report_codes()

    status_map: dict = {}
    try:
        rows = _pg_query(
            "select location_code, status, completion_pct from monthly_submissions "
            "where month_year = %s",
            (_mk_to_date(month_year),),
        )
        for r in rows:
            uid = r["location_code"]
            pct = float(r["completion_pct"] or 0)
            status_map[uid] = {
                "status":         _revert_if_deleted(uid, month_year, r["status"] or "NOT_STARTED", pct),
                "completion_pct": pct,
            }
    except Exception:
        pass

    results = []
    for loc in locs:
        if loc["userId"] in excluded:
            continue
        sd = status_map.get(loc["userId"],
                            {"status": "NOT_STARTED", "completion_pct": 0.0})
        results.append({**loc, **sd})
    return results


def create_revision_request(zone_id: str, location_id: str,
                            month_year: str, reason: str) -> dict:
    """Zone user raises a correction request for an already-submitted month.

    Postgres note: schema's status enum is 'PENDING'/'APPROVED'/'REJECTED'
    (not the Sheets version's 'PENDING_HQO') -- using the schema's own
    convention rather than fighting it. request_id is now just the row's
    real integer id as a string, not a random UUID fragment -- simpler,
    and still opaque to users since it's only ever round-tripped from a
    "list requests" call to an "act on this one" call, never typed by hand.
    Required a schema fix first: revision_requests had no `notes` column
    for storing a rejection reason -- added via ALTER TABLE (see schema.sql).
    """
    try:
        if not reason.strip():
            return {"ok": False, "msg": "Please provide a reason for the revision request."}

        mdate = _mk_to_date(month_year)
        existing = _pg_one(
            "select id from revision_requests where location_code = %s and month_year = %s "
            "and status = 'PENDING'",
            (location_id, mdate),
        )
        if existing:
            return {"ok": False,
                    "msg": "A revision request for this location/month is already pending HQO approval."}

        requester = _pg_one("select id from users where login_code = %s", (zone_id,))
        if not requester:
            return {"ok": False, "msg": f"Requesting user '{zone_id}' not found."}

        row = _pg_one(
            """
            insert into revision_requests (location_code, month_year, requested_by, reason, status)
            values (%s, %s, %s, %s, 'PENDING')
            returning id
            """,
            (location_id, mdate, requester["id"], reason),
        )
        req_id = str(row["id"])
        audit_log(zone_id, "RevisionRequest",
                  f"req={req_id} loc={location_id} month={month_year}")
        return {"ok": True, "msg": f"Revision request {req_id} submitted to HQO."}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def get_revision_requests(zone_filter: str = "") -> list:
    """Return revision requests; optionally filtered to a specific
    requesting user's login_code (matches the original's zone_filter,
    which was always the requesting Zone user's own id)."""
    try:
        sql = """
            select rr.id, rr.location_code, rr.month_year, rr.reason, rr.status,
                   rr.actioned_at, rr.notes, rr.created_at,
                   ru.login_code as requested_by_code,
                   au.login_code as actioned_by_code
            from revision_requests rr
            join users ru on ru.id = rr.requested_by
            left join users au on au.id = rr.actioned_by
        """
        params: tuple = ()
        if zone_filter:
            sql += " where ru.login_code = %s"
            params = (zone_filter,)
        sql += " order by rr.created_at desc"
        rows = _pg_query(sql, params)
        return [
            {
                "row":         r["id"],
                "request_id":  str(r["id"]),
                "zone_id":     r["requested_by_code"],
                "location_id": r["location_code"],
                "month_year":  month_key(r["month_year"]),
                "reason":      r["reason"] or "",
                "status":      r["status"] or "PENDING",
                "actioned_by": r["actioned_by_code"] or "",
                "actioned_at": r["actioned_at"].isoformat() if r["actioned_at"] else "",
                "notes":       r["notes"] or "",
                "created_at":  r["created_at"].isoformat() if r["created_at"] else "",
            }
            for r in rows
        ]
    except Exception:
        return []


def approve_revision_request(request_id: str, actioned_by: str) -> dict:
    """HQO approves revision — unlocks the location's month for re-editing."""
    try:
        row = _pg_one(
            "select location_code, month_year from revision_requests where id = %s",
            (int(request_id),),
        )
        if not row:
            return {"ok": False, "msg": f"Request ID {request_id} not found."}
        location_id = row["location_code"]
        month_year  = month_key(row["month_year"])

        actioner = _pg_one("select id from users where login_code = %s", (actioned_by,))
        _pg_query(
            "update revision_requests set status = 'APPROVED', actioned_by = %s, actioned_at = now() "
            "where id = %s",
            (actioner["id"] if actioner else None, int(request_id)), fetch=False,
        )

        sd = get_month_status(location_id, month_year)
        _update_submission_status(
            location_id, month_year, "REJECTED",
            sd.get("completion_pct", 0),
            checker_notes=(
                "Correction approved by HQO. "
                "Please update the data and resubmit."
            ),
        )
        audit_log(actioned_by, "ApproveRevision",
                  f"req={request_id} loc={location_id} month={month_year}")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def reject_revision_request(request_id: str,
                             actioned_by: str, note: str) -> dict:
    """HQO rejects revision request."""
    try:
        actioner = _pg_one("select id from users where login_code = %s", (actioned_by,))
        n = _pg_query(
            "update revision_requests set status = 'REJECTED', actioned_by = %s, "
            "actioned_at = now(), notes = %s where id = %s",
            (actioner["id"] if actioner else None, note, int(request_id)), fetch=False,
        )
        if not n:
            return {"ok": False, "msg": f"Request ID {request_id} not found."}
        audit_log(actioned_by, "RejectRevision", f"req={request_id} note={note[:60]}")
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


# ── Fix-1: Auto-populate Zone & HQO accounts ──────────────────────────────────

_UA_HEADERS = [
    "user_id", "location_name", "zone", "password", "role",
    "email", "mobile", "active",
]

def setup_zone_accounts() -> dict:
    """Auto-generate Zone and HQO accounts from existing zones/locations.

    Postgres note: zones are already their own rows (created during the
    Hackathon import), so this no longer needs to derive zone names from
    Maker rows -- it just ensures every zone has a Zone-role user, plus the
    HQO/Admin and view-only accounts.
    """
    try:
        zones = _pg_query("select id, name from zones")
        existing_zone_ids = {
            r["zone_id"] for r in _pg_query(
                "select zone_id from users where role = 'Zone' and zone_id is not null"
            )
        }
        added = []

        def _hash(pw: str) -> str:
            return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

        for z in zones:
            if z["id"] in existing_zone_ids:
                continue
            prefix   = z["name"][:3].upper()
            zone_uid = prefix + "ZONE"
            pw       = prefix + "MIS"
            _pg_query(
                """insert into users (login_code, zone_id, role, password_hash, is_first_login, active)
                   values (%s, %s, 'Zone', %s, true, true)
                   on conflict (login_code) do nothing""",
                (zone_uid, z["id"], _hash(pw)), fetch=False,
            )
            added.append(zone_uid)
            audit_log("SYSTEM", "SetupZoneAccount",
                      f"Added zone account {zone_uid} for zone {z['name']}")

        if not _pg_one("select 1 from users where login_code = 'SODSBU'"):
            _pg_query(
                """insert into users (login_code, role, password_hash, is_first_login, active)
                   values ('SODSBU', 'Admin', %s, true, true)""",
                (_hash("SODMIS"),), fetch=False,
            )
            added.append("SODSBU")
            audit_log("SYSTEM", "SetupZoneAccount", "Added HQO account SODSBU")

        if not _pg_one("select 1 from users where login_code = 'SODVIEW'"):
            _pg_query(
                """insert into users (login_code, role, password_hash, is_first_login, active)
                   values ('SODVIEW', 'Viewer', %s, true, true)""",
                (_hash("VIEWMIS"),), fetch=False,
            )
            added.append("SODVIEW")
            audit_log("SYSTEM", "SetupZoneAccount", "Added Viewer account SODVIEW")

        return {"ok": True, "added": added}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


# ── FLAGGED FOR REVIEW ────────────────────────────────────────────────────
# The four functions below (get_zone_admin_accounts, get_all_maker_credentials,
# get_all_checker_credentials, get_all_zone_credentials) let Admin view every
# user's PLAINTEXT password in the Sheets version -- used today for account
# review and for the credential-distribution emails (see app.py ~line 5629).
# Postgres stores bcrypt hashes (a real security improvement over the
# Sheets version), which are mathematically one-way: the original password
# CANNOT be recovered once hashed, by anyone, including this code.
# "password" is returned as None below so nothing crashes, but the actual
# credential-distribution workflow needs a product decision before cutover:
# most likely "generate a new password and show/email it once at creation
# time" rather than "view the existing one" -- worth discussing before this
# is relied on for real onboarding.

def get_zone_admin_accounts() -> list:
    rows = _pg_query("""
        select u.login_code, u.role, z.name as zone_name
        from users u left join zones z on z.id = u.zone_id
        where u.role in ('Zone', 'Admin')
    """)
    return [
        {"user_id": r["login_code"], "loc_name": r["zone_name"] or "",
         "zone": r["zone_name"] or "", "password": None, "role": r["role"]}
        for r in rows
    ]


def get_all_maker_credentials() -> list:
    rows = _pg_query("""
        select u.login_code, l.name as loc_name, z.name as zone_name
        from users u
        join locations l on l.code = u.location_code
        left join zones z on z.id = l.zone_id
        where u.role = 'Maker'
    """)
    return [
        {"userId": r["login_code"], "locName": r["loc_name"] or "",
         "zone": r["zone_name"] or "", "password": None}
        for r in rows
    ]


def get_all_checker_credentials() -> list:
    rows = _pg_query("""
        select u.login_code, l.name as loc_name, z.name as zone_name
        from users u
        join locations l on l.code = u.location_code
        left join zones z on z.id = l.zone_id
        where u.role = 'Checker'
    """)
    return [
        {"userId": r["login_code"], "locName": r["loc_name"] or "",
         "zone": r["zone_name"] or "", "password": None}
        for r in rows
    ]


def get_all_zone_credentials() -> list:
    """Deduplicates by zone name — keeps the first occurrence, matching the
    original's behavior to prevent double emails."""
    rows = _pg_query("""
        select u.login_code, z.name as zone_name
        from users u join zones z on z.id = u.zone_id
        where u.role = 'Zone'
        order by u.id
    """)
    out, seen = [], set()
    for r in rows:
        if r["zone_name"] in seen:
            continue
        seen.add(r["zone_name"])
        out.append({"userId": r["login_code"], "locName": r["zone_name"] or "",
                     "zone": r["zone_name"] or "", "password": None})
    return out

# ── end flagged section ────────────────────────────────────────────────────


def upsert_zone_account(zone_name: str, new_user_id: str, new_password: str) -> dict:
    """Create or update a Zone account matched by zone_name."""
    try:
        z = _pg_one("select id from zones where name = %s", (zone_name,))
        if not z:
            return {"ok": False, "msg": f"Zone '{zone_name}' not found."}
        pw_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()

        existing = _pg_one(
            "select id from users where role = 'Zone' and zone_id = %s", (z["id"],)
        )
        if existing:
            _pg_query(
                "update users set login_code = %s, password_hash = %s where id = %s",
                (new_user_id, pw_hash, existing["id"]), fetch=False,
            )
            audit_log("SYSTEM", "UpdateZoneAccount",
                      f"zone={zone_name} new_id={new_user_id}")
            return {"ok": True, "action": "updated"}

        _pg_query(
            """insert into users (login_code, zone_id, role, password_hash, is_first_login, active)
               values (%s, %s, 'Zone', %s, true, true)""",
            (new_user_id, z["id"], pw_hash), fetch=False,
        )
        audit_log("SYSTEM", "CreateZoneAccount", f"zone={zone_name} id={new_user_id}")
        return {"ok": True, "action": "created"}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def sync_missing_maker_accounts(default_password: str = "") -> dict:
    """Add Maker and Checker user rows for every location missing either role.

    FLAGGED FOR REVIEW: Checker's login_code here is location_code + "C"
    (e.g. "1775C") -- a distinct, unique login, matching the convention
    import_user_access.mjs already established during the Hackathon
    migration and confirmed present in the real data. This differs from the
    Sheets version, where Checker shares the SAME login_code as Maker and is
    distinguished only by which password matches during check_login's row
    scan. This is a real user-facing change: Checkers will need to know to
    type e.g. "1775C" instead of "1775" going forward. Worth confirming
    before cutover, and worth a heads-up to Checkers if so.
    """
    try:
        locations = _pg_query("select code, name from locations where active")
        existing = {(r["login_code"], r["role"])
                    for r in _pg_query("select login_code, role from users")}
        added, skipped = [], 0

        def _hash(pw: str) -> str:
            return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

        for loc in locations:
            code, name = loc["code"], loc["name"]
            base_pw = str(default_password).strip() or code

            if (code, "Maker") not in existing:
                _pg_query(
                    """insert into users (login_code, location_code, role, password_hash, is_first_login, active)
                       values (%s, %s, 'Maker', %s, true, true)""",
                    (code, code, _hash(base_pw)), fetch=False,
                )
                added.append(f"{code} (Maker)")
                audit_log("SYSTEM", "SyncAccount", f"Added Maker for {code} ({name})")
            else:
                skipped += 1

            checker_code = code + "C"
            if (checker_code, "Checker") not in existing:
                _pg_query(
                    """insert into users (login_code, location_code, role, password_hash, is_first_login, active)
                       values (%s, %s, 'Checker', %s, true, true)""",
                    (checker_code, code, _hash(base_pw + "C")), fetch=False,
                )
                added.append(f"{code} (Checker)")
                audit_log("SYSTEM", "SyncAccount", f"Added Checker for {code} ({name})")
            else:
                skipped += 1

        return {"ok": True, "added": added, "skipped": skipped}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def hqo_account_exists() -> bool:
    """Quick check: is there at least one Admin-role user?"""
    return bool(_pg_one("select 1 from users where role = 'Admin'"))


# ── App Settings (admin-controlled feature flags) ────────────────────────────

def register_session(user_id: str, token: str) -> dict:
    """Write session token for user. Uses the dedicated session columns on
    `users` (current_session_jti/current_session_started_at) instead of the
    original's Settings-sheet key/value hack -- the schema already has a
    proper place for this."""
    _pg_query(
        "update users set current_session_jti = %s, current_session_started_at = now() "
        "where login_code = %s",
        (token, user_id), fetch=False,
    )
    return {"ok": True}


def check_session_valid(user_id: str, token: str) -> bool:
    """Return True if stored session token matches the one in this session."""
    row = _pg_one(
        "select current_session_jti from users where login_code = %s", (user_id,)
    )
    return bool(row) and row["current_session_jti"] == token


def clear_session(user_id: str) -> None:
    """Remove the active session token for a user (on logout/timeout)."""
    _pg_query(
        "update users set current_session_jti = null, current_session_started_at = null "
        "where login_code = %s",
        (user_id,), fetch=False,
    )


def get_active_sessions(threshold_min: float = 30) -> list:
    """Return [(user_id, minutes_ago), ...] for session tokens younger than threshold_min.

    Approximates "who's currently logged in" from the session start timestamp
    written at login. A token older than the app's own 30-minute inactivity
    timeout is stale — that user has already timed out client-side even
    though the token record hasn't been cleared yet.
    """
    rows = _pg_query(
        """
        select login_code, current_session_started_at
        from users
        where current_session_jti is not null
          and current_session_started_at > now() - (%s || ' minutes')::interval
        """,
        (str(threshold_min),),
    )
    now = datetime.now(timezone.utc)
    result = []
    for r in rows:
        started = r["current_session_started_at"]
        if started is None:
            continue
        age_min = (now - started).total_seconds() / 60
        result.append((r["login_code"], age_min))
    return result


def get_setting(key: str, default: str = "FALSE") -> str:
    """Read a single value from app_settings."""
    try:
        row = _pg_one("select value from app_settings where key = %s", (key,))
        return row["value"] if row and row["value"] else default
    except Exception:
        return default


def set_setting(key: str, value: str, updated_by: str = "system") -> dict:
    """Write/update a value in app_settings."""
    try:
        _pg_query(
            """
            insert into app_settings (key, value, updated_by, updated_at)
            values (%s, %s, %s, now())
            on conflict (key) do update set
                value = excluded.value, updated_by = excluded.updated_by, updated_at = now()
            """,
            (key, value, updated_by), fetch=False,
        )
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


# ── Location management helpers ───────────────────────────────────────────────

def get_excluded_report_codes() -> set:
    """Return set of location codes excluded from pending-submission reports.

    Stored as comma-separated value under Settings key 'excluded_from_reports'.
    Excluded locations (e.g. non-operational) still appear in the app but are
    invisible to the Reports page, zone pending counts, and reminder emails.
    """
    val = get_setting("excluded_from_reports", "")
    return {c.strip() for c in val.split(",") if c.strip()}


def set_excluded_report_codes(codes: set, updated_by: str) -> dict:
    """Persist the exclusion list to Settings sheet."""
    val = ",".join(sorted(str(c) for c in codes))
    return set_setting("excluded_from_reports", val, updated_by)


def update_location_zone(loc_code: str, new_zone: str, updated_by: str) -> dict:
    """Update zone assignment for a location.

    Postgres note: locations.zone_id is the single source of truth (unlike
    the sheet, which denormalized zone onto every Maker row and needed a
    "update every matching row" loop) -- this updates exactly one row, and
    every user tied to that location reflects the change automatically via
    the join in check_login/_loc_name_map etc.
    """
    try:
        z = _pg_one("select id from zones where name = %s", (new_zone,))
        if not z:
            return {"ok": False, "msg": f"Zone '{new_zone}' not found."}
        n = _pg_query(
            "update locations set zone_id = %s where code = %s",
            (z["id"], str(loc_code)), fetch=False,
        )
        if n:
            _loc_name_map.clear()
            audit_log(updated_by, "ZoneUpdate", f"loc={loc_code} new_zone={new_zone}")
            return {"ok": True, "msg": f"Zone updated to '{new_zone}' for {loc_code}."}
        return {"ok": False, "msg": f"Location code {loc_code} not found."}
    except Exception as exc:
        return {"ok": False, "msg": str(exc)}


# ── EmailMaster: dynamic email maps ──────────────────────────────────────────

def get_email_master_maps() -> tuple:
    """Read email_routing → (loc_map, zone_map).

    Returns (None, None) if empty — callers fall back to the hardcoded
    dicts in emails.py.
    """
    try:
        rows = _pg_query("select entity_type, entity_code, email, cc from email_routing")
        if not rows:
            return None, None
        loc_map, zone_map = {}, {}
        for r in rows:
            if not r["entity_code"] or not r["email"]:
                continue
            if r["entity_type"] == "LOCATION":
                loc_map[r["entity_code"]] = r["email"]
            elif r["entity_type"] == "ZONE":
                zone_map[r["entity_code"]] = {"to": r["email"], "cc": r["cc"] or ""}
        return (loc_map or None), (zone_map or None)
    except Exception:
        return None, None


def seed_email_master(location_map: dict, zone_map: dict) -> dict:
    """Populate (or overwrite) email_routing from the given dicts.

    Called once from the Mail Trigger page to migrate hardcoded data so the
    admin can edit it in the database going forward.
    """
    try:
        _pg_query("delete from email_routing", fetch=False)
        count = 0
        for code, email in sorted(location_map.items()):
            _pg_query(
                "insert into email_routing (entity_type, entity_code, email) "
                "values ('LOCATION', %s, %s)",
                (code, email), fetch=False,
            )
            count += 1
        for zone, v in sorted(zone_map.items()):
            _pg_query(
                "insert into email_routing (entity_type, entity_code, display_name, email, cc) "
                "values ('ZONE', %s, %s, %s, %s)",
                (zone, zone, v.get("to", ""), v.get("cc", "")), fetch=False,
            )
            count += 1
        return {"ok": True, "count": count}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


# ── Phase-7: Excel Template Download & Upload ─────────────────────────────────

def generate_mis_template(
    user_id: str,
    month_year: str,
    user_info: dict,
    existing_draft: dict | None = None,
    loc_type: str = "HPCL",
) -> bytes:
    """Build a validated, protected .xlsx workbook with 4 sheets.

    Sheet 1 – MIS Data:
      Row 1  Section banners (HPCL gold, merged per section)
      Row 2  Field labels (HPCL blue; auto fields green)
      Row 3  Hints row (grey italic; includes unit / range / options)
      Row 4  Data entry — pre-filled from draft
             • Identity cols + auto-calc cols → LOCKED
             • User-input cols → UNLOCKED
             • Excel IFERROR formulas for all auto-calc fields
             • Data-validation rules (decimal / whole / list) per field
             • Input-message tooltip on every editable cell
             • Error alert (STOP) on every validated cell

    Sheets 2-4 – Railway Claims, IRR Details, Legal Cases:
      Header row locked; data rows unlocked, pre-filled from saved detail tables.

    Sheet protection password: HPCL@MIS (unlock in Excel if bulk edit needed).
    """
    import io as _io
    import re as _re
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, Protection
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from form_defs import SECTION_FIELDS, SECTION_NAMES

    PROTECT_PW = "HPCL@MIS"
    BIG_NUM    = "1E+15"          # effective "no upper limit"

    wb    = Workbook()
    draft = existing_draft or {}

    # ── Shared styles ────────────────────────────────────────────────────
    _thin   = Side(style="thin", color="CCCCCC")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _fill(hex6):
        return PatternFill("solid", fgColor=hex6)

    def _font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)

    BLUE_FILL  = _fill("0033A0")
    GOLD_FILL  = _fill("C6A64A")
    AUTO_FILL  = _fill("D6F5E0")   # light green — auto-calc
    LOCK_FILL  = _fill("F0F4FF")   # soft blue — locked identity
    HINT_FILL  = _fill("F8F9FA")
    WHITE_FILL = _fill("FFFFFF")
    NA_FILL    = _fill("D0D0D0")   # grey — N/A for this location type

    W_FONT  = _font(bold=True,  color="FFFFFF", size=9)
    AU_FONT = _font(italic=True, color="1a7a3c", size=9)
    HT_FONT = _font(italic=True, color="888888", size=8)
    NM_FONT = _font(size=10)
    ID_FONT = _font(bold=True,  color="0033A0", size=10)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    LOCKED_CELL   = Protection(locked=True)
    UNLOCKED_CELL = Protection(locked=False)

    def _cell(ws, row, col, value=None, font=None, fill=None,
              align=None, bdr=True, lock=True):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font       = font
        if fill:  c.fill       = fill
        if align: c.alignment  = align
        if bdr:   c.border     = _border
        c.protection = LOCKED_CELL if lock else UNLOCKED_CELL
        return c

    # ── Location-type exclusions ─────────────────────────────────────────
    from form_defs import get_excluded_fields as _get_excl
    _excl_keys = _get_excl(loc_type)   # frozenset of field keys not applicable

    # ── Collect all fields in section order ──────────────────────────────
    all_fields = []
    for sn in sorted(SECTION_FIELDS):
        for f in SECTION_FIELDS[sn]:
            all_fields.append((sn, f))

    ID_COLS = ["User ID", "Location Name", "Zone", "Month-Year"]
    ID_VALS = [user_id, user_info.get("locName", ""),
               user_info.get("zone", ""), month_year]
    N_ID    = len(ID_COLS)

    # ── Map field key → Excel column number (row 4) ──────────────────────
    field_col = {}
    for idx, (sn, f) in enumerate(all_fields):
        field_col[f["key"]] = N_ID + 1 + idx

    def _to_xl_formula(expr: str, row: int = 4) -> str:
        """Convert Python field expression (e.g. 'f17+f18+f19') to Excel formula."""
        def _sub(m):
            col = field_col.get(m.group(0))
            return f"{get_column_letter(col)}{row}" if col else "0"
        xl = _re.sub(r'f\d+', _sub, expr)
        return f'=IFERROR({xl},"")' if "/" in xl else f"={xl}"

    # ── Sheet 1: MIS Data ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "MIS Data"

    # Row 1 — Section banners (merged, HPCL gold)
    for ci, h in enumerate(ID_COLS, 1):
        _cell(ws1, 1, ci, h, font=W_FONT, fill=BLUE_FILL, align=CENTER)
        ws1.column_dimensions[get_column_letter(ci)].width = 16

    sec_col = N_ID + 1
    for sn in sorted(SECTION_FIELDS):
        fields = SECTION_FIELDS[sn]
        end_c  = sec_col + len(fields) - 1
        _cell(ws1, 1, sec_col, SECTION_NAMES[sn],
              font=_font(bold=True, color="FFFFFF", size=9),
              fill=GOLD_FILL, align=CENTER)
        if len(fields) > 1:
            ws1.merge_cells(start_row=1, start_column=sec_col,
                            end_row=1,   end_column=end_c)
        sec_col += len(fields)
    ws1.row_dimensions[1].height = 30

    # Row 2 — Field labels
    for ci, h in enumerate(ID_COLS, 1):
        _cell(ws1, 2, ci, h, font=W_FONT, fill=BLUE_FILL, align=CENTER)

    for ci, (sn, f) in enumerate(all_fields, N_ID + 1):
        is_auto = bool(f.get("auto"))
        is_na   = f["key"] in _excl_keys
        lbl = f["label"] + (" *" if f.get("req") and not is_auto else "")
        if is_auto:
            lbl += "  [Auto-Calc]"
        if is_na:
            lbl = f["label"] + "  [N/A — Not Applicable]"
        _cell(ws1, 2, ci, lbl,
              font=_font(italic=True, color="666666", size=9) if is_na
                   else (AU_FONT if is_auto else W_FONT),
              fill=NA_FILL if is_na else (AUTO_FILL if is_auto else BLUE_FILL),
              align=CENTER)
        ws1.column_dimensions[get_column_letter(ci)].width = 22
    ws1.row_dimensions[2].height = 60

    # Row 3 — Hints / validation notes
    for ci in range(1, N_ID + 1):
        _cell(ws1, 3, ci, "Pre-filled  —  do not edit",
              font=HT_FONT, fill=HINT_FILL, align=CENTER)

    for ci, (sn, f) in enumerate(all_fields, N_ID + 1):
        if f["key"] in _excl_keys:
            _cell(ws1, 3, ci, f"Not applicable for {loc_type} locations — leave blank",
                  font=HT_FONT, fill=NA_FILL, align=CENTER)
            continue
        parts = [f.get("hint", "")]
        if f.get("min") is not None:
            parts.append(f"Min: {f['min']}")
        if f.get("max") is not None:
            parts.append(f"Max: {f['max']}")
        if f["type"] == "number" and f.get("dec") is not None:
            parts.append(f"Decimals: up to {f['dec']}")
        if f["type"] == "int":
            parts.append("Whole numbers only")
        if f.get("opts"):
            parts.append("Options: " + " / ".join(f["opts"]))
        if f.get("auto"):
            parts.append("[Auto-calculated — do not enter manually]")
        if f["type"] == "textarea":
            parts.append("Press Alt+Enter to add a new line")
        hint_text = "  |  ".join(p for p in parts if p)
        _cell(ws1, 3, ci, hint_text, font=HT_FONT, fill=HINT_FILL, align=LEFT)
    ws1.row_dimensions[3].height = 30

    # Row 4 — Data (pre-filled + protected + validated)
    for ci, (h, v) in enumerate(zip(ID_COLS, ID_VALS), 1):
        _cell(ws1, 4, ci, v, font=ID_FONT, fill=LOCK_FILL, align=CENTER, lock=True)

    for ci, (sn, f) in enumerate(all_fields, N_ID + 1):
        is_auto   = bool(f.get("auto"))
        is_na     = f["key"] in _excl_keys
        raw       = draft.get(f["key"])
        cell_ref  = f"{get_column_letter(ci)}4"

        if is_na:
            _cell(ws1, 4, ci, "N/A",
                  font=_font(italic=True, color="888888", size=10),
                  fill=NA_FILL, align=CENTER, lock=True)
            continue

        if is_auto:
            formula = _to_xl_formula(f["auto"])
            c = _cell(ws1, 4, ci, formula,
                      font=AU_FONT, fill=AUTO_FILL, align=CENTER, lock=True)
            # Apply decimal number format so auto-calc results display correctly
            _dec = f.get("dec") or 2
            c.number_format = "0." + "0" * _dec
        else:
            val = None
            if raw not in (None, ""):
                try:
                    val = float(raw) if f["type"] in ("number", "int") else str(raw)
                except Exception:
                    val = str(raw)
            c = _cell(ws1, 4, ci, val,
                      font=NM_FONT, fill=WHITE_FILL, align=CENTER, lock=False)

            # ── Data validation ─────────────────────────────────────────
            ftype = f["type"]
            mn    = f.get("min")
            mx    = f.get("max")
            opts  = f.get("opts")
            dec   = f.get("dec")
            lbl   = f["label"]
            hint  = f.get("hint", "")

            # Build a descriptive range string for messages
            range_txt = ""
            if mn is not None and mx is not None:
                range_txt = f"between {mn} and {mx}"
            elif mn is not None:
                range_txt = f"≥ {mn}"
            elif mx is not None:
                range_txt = f"≤ {mx}"

            if ftype in ("number", "int") and (mn is not None or mx is not None):
                xl_type  = "whole" if ftype == "int" else "decimal"
                f1       = str(mn) if mn is not None else "0"
                f2       = str(mx) if mx is not None else BIG_NUM

                if mn is not None and mx is not None:
                    op = "between"
                elif mn is not None:
                    op, f2 = "greaterThanOrEqual", None
                else:
                    op, f1 = "lessThanOrEqual", f2

                dv_kwargs = dict(
                    type=xl_type, operator=op,
                    formula1=f1,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorStyle="stop",
                    errorTitle="Invalid Value",
                    error=(
                        f'"{lbl}" must be a '
                        f'{"whole number" if ftype=="int" else "decimal number"}'
                        + (f" {range_txt}" if range_txt else "")
                        + ("." if not range_txt else "")
                        + (f" (up to {dec} decimal places)" if ftype == "number" and dec is not None else "")
                    ),
                    showInputMessage=True,
                    promptTitle=lbl[:32],
                    prompt=(
                        hint
                        + (f"\nRange: {range_txt}" if range_txt else "")
                        + (f"\nDecimals: up to {dec}" if ftype == "number" and dec is not None else "")
                        + ("\nWhole numbers only" if ftype == "int" else "")
                    ),
                )
                if f2 is not None:
                    dv_kwargs["formula2"] = f2

                dv = DataValidation(**dv_kwargs)
                dv.sqref = cell_ref
                ws1.add_data_validation(dv)

            elif ftype == "select" and opts:
                opts_str = ",".join(opts)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{opts_str}"',
                    allow_blank=True,
                    showErrorMessage=True,
                    errorStyle="stop",
                    errorTitle="Invalid Selection",
                    error=f'Select one of: {opts_str}',
                    showInputMessage=True,
                    promptTitle=lbl[:32],
                    prompt=f"{hint}\nSelect: {opts_str}",
                )
                dv.sqref = cell_ref
                ws1.add_data_validation(dv)

            elif ftype == "date":
                # Store as text in DD/MM/YYYY format; enforce format via custom validation
                c.number_format = "@"  # force text so Excel doesn't auto-convert to serial
                _cr0 = cell_ref       # e.g. "AZ4"
                formula = (
                    f'=AND(LEN({_cr0})=10,'
                    f'ISNUMBER(VALUE(LEFT({_cr0},2))),'
                    f'MID({_cr0},3,1)="/",'
                    f'ISNUMBER(VALUE(MID({_cr0},4,2))),'
                    f'MID({_cr0},6,1)="/",'
                    f'ISNUMBER(VALUE(RIGHT({_cr0},4))),'
                    f'VALUE(LEFT({_cr0},2))>=1,'
                    f'VALUE(LEFT({_cr0},2))<=31,'
                    f'VALUE(MID({_cr0},4,2))>=1,'
                    f'VALUE(MID({_cr0},4,2))<=12)'
                )
                dv = DataValidation(
                    type="custom", formula1=formula,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorStyle="warning",
                    errorTitle="Invalid Date Format",
                    error='Enter date as DD/MM/YYYY — e.g. 25/06/2026',
                    showInputMessage=True,
                    promptTitle=lbl[:32],
                    prompt=f"{hint}\nFormat: DD/MM/YYYY (e.g. 25/06/2026)",
                )
                dv.sqref = cell_ref
                ws1.add_data_validation(dv)

            else:
                # textarea / unconstrained — hint is in row 3; no validation rule needed
                pass

    # Taller data row so textarea cells (multi-line text) are readable
    ws1.row_dimensions[4].height = 60
    ws1.freeze_panes = ws1.cell(row=4, column=N_ID + 1)

    # ── Protect MIS Data sheet ───────────────────────────────────────────
    ws1.protection.sheet             = True
    ws1.protection.password          = PROTECT_PW
    ws1.protection.selectLockedCells   = False   # allow clicking locked cells
    ws1.protection.selectUnlockedCells = False

    # Date field keys per detail tab — these get DD/MM/YYYY formula validation
    _DETAIL_DATE_KEYS = {
        "RAILWAY_CLAIMS": {"last_hearing", "next_hearing"},
        "IRR_DETAILS":    {"irr_date", "closure_date"},
        "LEGAL_CASES":    {"last_hearing", "next_hearing"},
    }

    # ── Helper: protected detail sheet ──────────────────────────────────
    def _detail_sheet(sheet_name: str, tab_key: str):
        ddef      = _DETAIL_DEF[tab_key]
        data_keys = ddef["data_keys"]
        col_hdrs  = ddef["sheet_headers"][ddef["prefix_count"]:]
        existing  = load_detail_table(user_id, month_year, tab_key)
        date_keys = _DETAIL_DATE_KEYS.get(tab_key, set())

        ws2 = wb.create_sheet(sheet_name)

        # Header row — locked
        for ci, lbl in enumerate(col_hdrs, 1):
            _cell(ws2, 1, ci, lbl, font=W_FONT, fill=BLUE_FILL, align=CENTER, lock=True)
            ws2.column_dimensions[get_column_letter(ci)].width = 22
        ws2.row_dimensions[1].height = 30

        # Data rows — unlocked
        data_rows = existing if existing else [{}]
        for ri, rec in enumerate(data_rows, 2):
            for ci, key in enumerate(data_keys, 1):
                val = rec.get(key) or None
                _cell(ws2, ri, ci, val,
                      font=NM_FONT, fill=WHITE_FILL, align=LEFT, lock=False)
            ws2.row_dimensions[ri].height = 20

        # Extend blank unlocked rows to row 200
        start_blank = 2 + len(data_rows)
        for ri in range(start_blank, 201):
            for ci in range(1, len(col_hdrs) + 1):
                ws2.cell(row=ri, column=ci).protection = UNLOCKED_CELL

        # Data validation per column — DD/MM/YYYY for dates, input-message for others
        for ci, (lbl, key) in enumerate(zip(col_hdrs, data_keys), 1):
            cl      = get_column_letter(ci)
            col_ref = f"{cl}2:{cl}200"
            if key in date_keys:
                cell0   = f"{cl}2"
                formula = (f'=AND(LEN({cell0})=10,'
                           f'MID({cell0},3,1)="/",'
                           f'MID({cell0},6,1)="/")')
                dv = DataValidation(
                    type="custom", formula1=formula,
                    allow_blank=True, showErrorMessage=True,
                    errorStyle="warning", errorTitle="Invalid Date Format",
                    error='Enter date as DD/MM/YYYY (e.g. 25/06/2025) or "NA" if unknown.',
                    showInputMessage=True, promptTitle=lbl[:32],
                    prompt="DD/MM/YYYY — e.g. 25/06/2025  (or NA)")
            else:
                dv = DataValidation(
                    allow_blank=True, showInputMessage=True,
                    promptTitle=lbl[:32],
                    prompt=f'Enter: {lbl}  (or NA if not applicable)',
                    showErrorMessage=False)
            dv.sqref = col_ref
            ws2.add_data_validation(dv)

        ws2.freeze_panes = ws2["A2"]

        ws2.protection.sheet             = True
        ws2.protection.password          = PROTECT_PW
        ws2.protection.selectLockedCells   = False
        ws2.protection.selectUnlockedCells = False

    _detail_sheet("Railway Claims", "RAILWAY_CLAIMS")
    _detail_sheet("IRR Details",    "IRR_DETAILS")
    _detail_sheet("Legal Cases",    "LEGAL_CASES")

    # ── S5A: all 10 M&I MIS subsection sheets ────────────────────────────
    # Each entry: (tab_key, sheet_name, headers, hints, keys, dropdowns)
    # dropdowns: dict of key → "opt1,opt2,..." for list-validation columns

    # Build tank dropdown for tank_no columns — inline list (reliable) with
    # hidden-sheet fallback for locations that have more than ~30 tanks.
    _loc_tanks    = get_tank_master().get(user_id, [])
    _loc_tanks_all = _loc_tanks + ["Other Tanks"]
    _tank_inline  = ",".join(_loc_tanks_all)          # e.g. "T-001,T-002,Other Tanks"
    _use_inline   = len(_tank_inline) <= 250           # Excel list-validation limit
    _n_tanks      = len(_loc_tanks_all)
    # Hidden TankList sheet — used when inline exceeds 250 chars
    ws_tl = wb.create_sheet("TankList")
    ws_tl.sheet_state = "hidden"
    for _ti, _tn in enumerate(_loc_tanks_all, 1):
        ws_tl.cell(row=_ti, column=1, value=_tn)

    _MI_TAB_DEFS = [
        ("MI_TANK_OUTAGE", "S5A-1 Tank Outage",
         ["Tank No.", "Other Tank Desc.", "Planned Start", "Planned End",
          "Actual Start", "Actual End", "Outage For", "Current Status"],
         ["Select tank number", "Describe if 'Other Tanks'",
          "DD/MM/YYYY", "DD/MM/YYYY", "DD/MM/YYYY", "DD/MM/YYYY",
          "Reason for outage", "Current status of outage"],
         ["tank_no", "other_tank_desc", "planned_start", "planned_end",
          "actual_start", "actual_end", "outage_for", "current_status"],
         {}),

        ("MI_MAJOR_REPAIR", "S5A-2 Major Repair",
         ["Tank No.", "Other Tank Desc.", "Nature of Repair",
          "Revenue / Capex", "AR Code", "Status", "ETC Date"],
         ["Select tank number", "Describe if 'Other Tanks'",
          "Describe nature of repair", "Select: Revenue or Capex",
          "AR code if applicable", "Current repair status", "DD/MM/YYYY — Expected completion"],
         ["tank_no", "other_tank_desc", "nature_of_repair",
          "revenue_capex", "ar_code", "current_status", "etc_date"],
         {"revenue_capex": "Revenue,Capex"}),

        ("MI_VRU", "S5A-3 VRU",
         ["VRU Operational", "Date Not Operating", "Action Taken", "ETC Date",
          "MS Vol Recovered (KL)", "Inlet MFM Start (m³)", "Inlet MFM End (m³)",
          "Outlet MFM Start (m³)", "Outlet MFM End (m³)", "Vapour Treated (m³)",
          "VOC Value (mg/cc)", "Inlet Emission (mg/cc)",
          "MS/Gasohol TT Vol (KL)", "HSD TT Vol (KL)",
          "MS/Gasohol TW Vol (KL)", "HSD TW Vol (KL)", "VRU Uptime (%)"],
         ["Yes/No", "DD/MM/YYYY if not operating", "Action taken if not operating",
          "DD/MM/YYYY — Expected completion", "Numeric",
          "Numeric (m³)", "Numeric (m³)", "Numeric (m³)", "Numeric (m³)",
          "Numeric (m³)", "VOC concentration at VRU outlet mg/cc",
          "VOC concentration at VRU inlet mg/cc",
          "Numeric (KL)", "Numeric (KL)", "Numeric (KL)", "Numeric (KL)", "0–100"],
         ["vru_operational", "date_not_operating", "action_taken", "etc_date",
          "ms_vol_recovered_kl", "inlet_mfm_start_m3", "inlet_mfm_end_m3",
          "outlet_mfm_start_m3", "outlet_mfm_end_m3", "vapour_treated_m3",
          "voc_value_mgcc", "inlet_emission_mgcc",
          "ms_gasohol_tt_vol_kl", "hsd_tt_vol_kl",
          "ms_gasohol_tw_vol_kl", "hsd_tw_vol_kl", "vru_uptime_pct"],
         {"vru_operational": "Yes,No"}),

        ("MI_AUDIT_2526", "S5A-4 M&I Audit 25-26",
         ["Audit Date", "No. of Recommendations", "No. Pending", "External Score"],
         ["DD/MM/YYYY", "Total recommendations from audit",
          "Pending recommendations", "Score from external auditor"],
         ["audit_date", "no_recommendations", "no_pending", "external_score"],
         {}),

        ("MI_AUDIT_2627", "S5A-5 M&I Audit 26-27",
         ["Audit Carried Out", "Audit Date", "No. of Recommendations",
          "No. Pending", "External Score"],
         ["Yes/No", "DD/MM/YYYY", "Total recommendations",
          "Pending recommendations", "Score from external auditor"],
         ["audit_carried_out", "audit_date", "no_recommendations",
          "no_pending", "external_score"],
         {"audit_carried_out": "Yes,No"}),

        ("MI_TECH_AUDIT", "S5A-6 Tech. Audit",
         ["Audit Date", "No. of Recommendations", "No. Pending", "Ref. No."],
         ["DD/MM/YYYY", "Total recommendations", "Pending count", "Reference number"],
         ["audit_date", "no_recommendations", "no_pending", "ref_no"],
         {}),

        ("MI_EQUIP_BREAKDOWN", "S5A-7 Equip. Breakdown",
         ["Equipment Name", "Equipment Other", "Equipment Details",
          "Start Date", "Issue", "Proposed Date", "Actual End Date",
          "Resolution Action"],
         ["Select equipment type", "Specify if 'Other'", "Details of breakdown",
          "DD/MM/YYYY", "Describe the issue",
          "DD/MM/YYYY — Proposed fix date", "DD/MM/YYYY — Actual resolution",
          "Action taken to resolve"],
         ["equipment_name", "equipment_other", "equipment_details",
          "start_date", "issue", "proposed_date", "actual_end_date",
          "resolution_action"],
         {"equipment_name": "Pipeline,Pump,Fire Fighting Equipment,Fire Engine,DG Set,Other"}),

        ("MI_INT_PIPELINE", "S5A-8 Int. Pipeline",
         ["Last UT Date", "Last Hydrotest Date", "Last DCVG Date",
          "Last LRUT Date", "Other Testing"],
         ["DD/MM/YYYY", "DD/MM/YYYY", "DD/MM/YYYY", "DD/MM/YYYY",
          "Describe any other testing done"],
         ["last_ut_date", "last_hydrotest_date", "last_dcvg_date",
          "last_lrut_date", "other_testing"],
         {}),

        ("MI_EXT_PIPELINE", "S5A-9 Ext. Pipeline",
         ["Pipeline Type", "Pipeline Details", "Length Metres", "Product", "Size Inch",
          "Last UT Date", "Last Hydrotest Date", "Last DCVG Date",
          "Last LRUT Date", "Other Testing"],
         ["UG = Underground / AG = Above Ground",
          "Describe pipeline segment (route / from-to)", "Length in metres",
          "Product carried e.g. MS HSD ATF",
          "Nominal bore in inches", "DD/MM/YYYY", "DD/MM/YYYY", "DD/MM/YYYY",
          "DD/MM/YYYY", "Describe any other testing"],
         ["pipeline_type", "pipeline_details", "length_metres", "product", "size_inch",
          "last_ut_date", "last_hydrotest_date", "last_dcvg_date",
          "last_lrut_date", "other_testing"],
         {"pipeline_type": "UG,AG"}),

        ("MI_TANK_STATUS", "S5A-10 Tank Status",
         ["Tank No", "Cleaning Completed Date", "Cleaning Due Date",
          "Extension Taken", "Extension EFN No",
          "Inspection Date", "Inspection Due Date",
          "Painting Date", "Painting Due Date",
          "Tank Status", "Tank Status Other"],
         ["Select tank number from Tank Master",
          "DD/MM/YYYY", "DD/MM/YYYY",
          "Yes / No / NA", "Required if Extension = Yes",
          "DD/MM/YYYY", "DD/MM/YYYY",
          "DD/MM/YYYY", "DD/MM/YYYY",
          "Operational / Under Repair / Under Cleaning / Idle / Revamp / Others",
          "Required if Tank Status = Others"],
         ["tank_no",
          "cleaning_completed_date", "cleaning_due_date",
          "extension_taken", "extension_efn_no",
          "inspection_date", "inspection_due_date",
          "painting_date", "painting_due_date",
          "tank_status", "tank_status_other"],
         {"extension_taken": "Yes,No,NA",
          "tank_status":     "Operational,Under Repair,Under Cleaning,Idle,Revamp,Others"}),
    ]

    BANNER_FILL = _fill("1a1a6e")
    BANNER_FONT = _font(bold=True, color="FFFFFF", size=11)
    NA_FONT     = _font(italic=True, color="888888", size=10)
    NA_FILL     = _fill("F5F5F5")

    # Date field keys per M&I tab — these get DD/MM/YYYY formula validation
    _MI_DATE_KEYS: dict[str, set] = {
        "MI_TANK_OUTAGE":     {"planned_start", "planned_end", "actual_start", "actual_end"},
        "MI_MAJOR_REPAIR":    {"etc_date"},
        "MI_VRU":             {"date_not_operating", "etc_date"},
        "MI_AUDIT_2526":      {"audit_date"},
        "MI_AUDIT_2627":      {"audit_date"},
        "MI_TECH_AUDIT":      {"audit_date"},
        "MI_EQUIP_BREAKDOWN": {"start_date", "proposed_date", "actual_end_date"},
        "MI_INT_PIPELINE":    {"last_ut_date", "last_hydrotest_date",
                               "last_dcvg_date", "last_lrut_date"},
        "MI_EXT_PIPELINE":    {"last_ut_date", "last_hydrotest_date",
                               "last_dcvg_date", "last_lrut_date"},
        "MI_TANK_STATUS":     {"cleaning_completed_date", "cleaning_due_date",
                               "inspection_date", "inspection_due_date",
                               "painting_date", "painting_due_date"},
    }

    for tab_key, sheet_name, hdrs, hints, keys, dropdowns in _MI_TAB_DEFS:
        ws_mi  = wb.create_sheet(sheet_name)
        n_cols = len(hdrs)
        date_keys_tab = _MI_DATE_KEYS.get(tab_key, set())

        # Row 1 — banner
        _cell(ws_mi, 1, 1, sheet_name,
              font=BANNER_FONT, fill=BANNER_FILL, align=CENTER)
        ws_mi.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws_mi.row_dimensions[1].height = 28

        # Row 2 — headers
        for ci, h in enumerate(hdrs, 1):
            _cell(ws_mi, 2, ci, h, font=W_FONT, fill=BLUE_FILL, align=CENTER, lock=True)
            ws_mi.column_dimensions[get_column_letter(ci)].width = 24
        ws_mi.row_dimensions[2].height = 40

        # Row 3 — hints (date fields show DD/MM/YYYY prominently)
        for ci, (h, key) in enumerate(zip(hints, keys), 1):
            display_hint = (f"DD/MM/YYYY — {h}" if key in date_keys_tab
                            else ("NA if not applicable — " + h if not h.startswith("NA") else h))
            _cell(ws_mi, 3, ci, display_hint, font=HT_FONT, fill=HINT_FILL, align=LEFT, lock=True)
        ws_mi.row_dimensions[3].height = 22

        # Pre-fill saved data rows starting at row 4
        existing   = load_mi_data(tab_key, user_id, month_year)
        is_na_tab  = bool(existing and existing[0].get("na_flag") == "Y")
        saved_rows = ([r for r in existing if r.get("na_flag") != "Y"]
                      if existing and not is_na_tab else [])

        if is_na_tab:
            # Write "Not Applicable" in first cell so users know tab is marked NA
            c = ws_mi.cell(row=4, column=1, value="Not Applicable — marked NA in application")
            c.font = NA_FONT; c.fill = NA_FILL; c.alignment = LEFT
            ws_mi.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
            ws_mi.row_dimensions[4].height = 20
            start_blank = 5
        else:
            for ri, rec in enumerate(saved_rows, 4):
                for ci, key in enumerate(keys, 1):
                    val = rec.get(key) or None
                    _cell(ws_mi, ri, ci, val, font=NM_FONT, fill=WHITE_FILL,
                          align=LEFT, lock=False)
                    if key in date_keys_tab:
                        ws_mi.cell(row=ri, column=ci).number_format = "@"
                ws_mi.row_dimensions[ri].height = 20
            start_blank = 4 + len(saved_rows)

        # Blank unlocked rows to row 150.
        # Use _cell() (not bare .protection) so protection + border are registered
        # together in a single openpyxl style entry, ensuring applyProtection="1"
        # is written to styles.xml and Excel honours the unlocked state.
        for ri in range(start_blank, 151):
            for ci in range(1, n_cols + 1):
                c = _cell(ws_mi, ri, ci, value=None,
                          font=NM_FONT, fill=WHITE_FILL, align=LEFT, lock=False)
                if keys[ci - 1] in date_keys_tab:
                    c.number_format = "@"

        # Data validations per column (DD/MM/YYYY formula for dates; list for dropdowns)
        for ci, (hdr, hint, key) in enumerate(zip(hdrs, hints, keys), 1):
            cl      = get_column_letter(ci)
            col_ref = f"{cl}4:{cl}150"
            if key in date_keys_tab:
                # Custom formula: check DD/MM/YYYY text format
                cell0   = f"{cl}4"
                formula = (f'=AND(LEN({cell0})=10,'
                           f'MID({cell0},3,1)="/",'
                           f'MID({cell0},6,1)="/")')
                dv = DataValidation(
                    type="custom", formula1=formula,
                    allow_blank=True, showErrorMessage=True,
                    errorStyle="warning", errorTitle="Invalid Date Format",
                    error='Enter date as DD/MM/YYYY (e.g. 25/06/2025) or "NA" if not applicable.',
                    showInputMessage=True, promptTitle=hdr[:32],
                    prompt=f"DD/MM/YYYY (e.g. 25/06/2025) or NA")
            elif key == "tank_no":
                _f1 = (f'"{_tank_inline}"' if _use_inline
                       else f"'TankList'!$A$1:$A${_n_tanks}")
                dv = DataValidation(
                    type="list", formula1=_f1,
                    allow_blank=True, showErrorMessage=True,
                    errorStyle="warning",
                    error="Select a tank from the list, or type the tank number directly if not listed.",
                    showInputMessage=True, promptTitle="Tank No.",
                    prompt="Select from list — OR type tank number directly if not listed (e.g. T-999).")
            elif key in dropdowns:
                dv = DataValidation(
                    type="list", formula1=f'"{dropdowns[key]}"',
                    allow_blank=True, showErrorMessage=True,
                    errorStyle="warning", error=f"Select: {dropdowns[key]}",
                    showInputMessage=True, promptTitle=hdr[:32], prompt=hint)
            else:
                dv = DataValidation(
                    allow_blank=True, showInputMessage=True,
                    promptTitle=hdr[:32],
                    prompt=f'Enter value, or "NA" if not applicable.',
                    showErrorMessage=False)
            dv.sqref = col_ref
            ws_mi.add_data_validation(dv)

        ws_mi.freeze_panes = ws_mi["A4"]
        ws_mi.protection.sheet           = True
        ws_mi.protection.password        = PROTECT_PW
        ws_mi.protection.selectLockedCells   = True   # prevent clicks on locked header rows
        ws_mi.protection.selectUnlockedCells = False  # allow selecting/editing data rows

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_filled_mis_report(
    user_id: str,
    month_year: str,
    user_info: dict,
    existing_draft: dict | None = None,
) -> bytes:
    """Generate a read-only filled MIS report Excel — all cells locked.

    Intended for download AFTER Checker approval (status = SUBMITTED).
    Same layout as generate_mis_template but fully locked and pre-filled.
    """
    import io as _io
    import re as _re
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from form_defs import SECTION_FIELDS, SECTION_NAMES

    draft = existing_draft or {}
    wb    = Workbook()

    _thin   = Side(style="thin", color="CCCCCC")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _fill(h):
        return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)

    BLUE_FILL  = _fill("0033A0")
    GOLD_FILL  = _fill("C6A64A")
    AUTO_FILL  = _fill("E8F5E9")
    HINT_FILL  = _fill("F8F9FA")
    WHITE_FILL = _fill("FFFFFF")
    W_FONT     = _font(bold=True, color="FFFFFF", size=9)
    HT_FONT    = _font(italic=True, color="888888", size=8)
    NM_FONT    = _font(size=10)
    AU_FONT    = _font(italic=True, color="1a7a3c", size=9)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    LOCKED     = Protection(locked=True)

    def _cell(ws, row, col, value=None, font=None, fill=None, align=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font      = font
        if fill:  c.fill      = fill
        if align: c.alignment = align
        c.border     = _border
        c.protection = LOCKED
        return c

    all_fields = []
    for sn in sorted(SECTION_FIELDS):
        for f in SECTION_FIELDS[sn]:
            all_fields.append((sn, f))

    ID_COLS = ["User ID", "Location Name", "Zone", "Month-Year"]
    ID_VALS = [user_id, user_info.get("locName", ""),
               user_info.get("zone", ""), month_year]
    N_ID    = len(ID_COLS)

    field_col = {f["key"]: N_ID + 1 + idx for idx, (_, f) in enumerate(all_fields)}

    def _to_xl(expr, row=4):
        def _sub(m):
            col = field_col.get(m.group(0))
            return f"{get_column_letter(col)}{row}" if col else "0"
        xl = _re.sub(r'f\d+', _sub, expr)
        return f'=IFERROR({xl},"")' if "/" in xl else f"={xl}"

    ws1 = wb.active
    ws1.title = "MIS Report"

    # Row 1 — report banner
    total_cols = N_ID + len(all_fields)
    _cell(ws1, 1, 1,
          f"HPCL SOD MIS REPORT — {user_info.get('locName','')} | {month_year}",
          font=_font(bold=True, color="FFFFFF", size=11),
          fill=_fill("001060"), align=CENTER)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 50))
    ws1.row_dimensions[1].height = 28

    # Row 2 — Section banners
    for ci, h in enumerate(ID_COLS, 1):
        _cell(ws1, 2, ci, h, font=W_FONT, fill=BLUE_FILL, align=CENTER)
        ws1.column_dimensions[get_column_letter(ci)].width = 16

    sec_col = N_ID + 1
    for sn in sorted(SECTION_FIELDS):
        fields = SECTION_FIELDS[sn]
        end_c  = sec_col + len(fields) - 1
        _cell(ws1, 2, sec_col, SECTION_NAMES[sn],
              font=_font(bold=True, color="FFFFFF", size=9),
              fill=GOLD_FILL, align=CENTER)
        if len(fields) > 1:
            ws1.merge_cells(start_row=2, start_column=sec_col,
                            end_row=2,   end_column=end_c)
        sec_col += len(fields)
    ws1.row_dimensions[2].height = 28

    # Row 3 — Field labels
    for ci, h in enumerate(ID_COLS, 1):
        _cell(ws1, 3, ci, h, font=W_FONT, fill=BLUE_FILL, align=CENTER)

    for ci, (sn, f) in enumerate(all_fields, N_ID + 1):
        is_auto = bool(f.get("auto"))
        lbl     = f["label"] + ("  [Auto]" if is_auto else "")
        _cell(ws1, 3, ci, lbl,
              font=AU_FONT if is_auto else W_FONT,
              fill=AUTO_FILL if is_auto else BLUE_FILL,
              align=CENTER)
        ws1.column_dimensions[get_column_letter(ci)].width = 18
    ws1.row_dimensions[3].height = 30

    # Row 4 — Hint row
    for ci, _ in enumerate(ID_COLS, 1):
        _cell(ws1, 4, ci, "", font=HT_FONT, fill=HINT_FILL, align=LEFT)
    for ci, (_, f) in enumerate(all_fields, N_ID + 1):
        hint = f.get("hint", "")
        _cell(ws1, 4, ci, hint, font=HT_FONT, fill=HINT_FILL, align=LEFT)
    ws1.row_dimensions[4].height = 20

    # Row 5 — Data values
    for ci, v in enumerate(ID_VALS, 1):
        _cell(ws1, 5, ci, v, font=_font(bold=True, color="0033A0", size=10),
              fill=_fill("F0F4FF"), align=LEFT)

    for ci, (sn, f) in enumerate(all_fields, N_ID + 1):
        if f.get("auto") and f.get("auto"):
            val = _to_xl(f["auto"], row=5)
        else:
            raw = draft.get(f["key"], "")
            val = raw if raw not in ("", None) else ""
        _cell(ws1, 5, ci, val, font=NM_FONT, fill=WHITE_FILL, align=LEFT)
    ws1.row_dimensions[5].height = 22

    ws1.freeze_panes = ws1["A5"]

    ws1.protection.sheet             = True
    ws1.protection.password          = "HPCL@MIS"
    ws1.protection.selectLockedCells   = False
    ws1.protection.selectUnlockedCells = False

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_mis_pdf_report(
    user_id: str,
    month_year: str,
    user_info: dict,
    draft: dict,
    status: str = "",
) -> bytes:
    """Generate a printer-friendly PDF review report for the Checker.

    Renders all MIS sections (S1-S10) with field labels and filled values,
    plus a summary of S5A M&I MIS tabs. Returns raw PDF bytes.
    """
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
    from form_defs import SECTION_FIELDS, SECTION_NAMES

    _SECTIONS = [
        (1, "Operations"),         (2, "Facilities & Planning"),
        (3, "S&D"),                (4, "Biofuel"),
        (5, "M&I"),                (6, "HSE"),
        (7, "Operational Efficiency"), (8, "EM Lock"),
        (9, "Transportation"),     (10, "Others"),
    ]
    _STATUS_LABELS = {
        "PENDING_REVIEW": "Pending Checker Review",
        "SUBMITTED":      "Approved & Locked",
        "IN_PROGRESS":    "In Progress",
        "REJECTED":       "Rejected",
    }

    BLUE   = (0,   51,  160)
    LBLUE  = (224, 232, 255)
    DARK   = (25,  25,  25)
    MGREY  = (100, 100, 100)
    LGREY  = (248, 248, 248)
    WHITE  = (255, 255, 255)
    GREEN  = (220, 245, 225)

    loc_name  = user_info.get("locName", user_id)
    zone_name = user_info.get("zone", "")

    def _s(text):
        """Sanitise to latin-1 for fpdf2 core fonts."""
        if text is None:
            return ""
        return str(text).encode("latin-1", "replace").decode("latin-1")

    def _val(v):
        return _s(v) if v not in (None, "", "None") else "-"

    class _PDF(FPDF):
        def header(self):
            self.set_fill_color(*BLUE)
            self.rect(0, 0, 210, 12, "F")
            self.set_font("Helvetica", "B", 9)
            self.set_text_color(*WHITE)
            self.set_xy(10, 2)
            self.cell(130, 8, "HPCL SOD e-MIS  |  Checker Review Report",
                      new_x=XPos.RIGHT, new_y=YPos.TOP)
            self.set_font("Helvetica", "", 7.5)
            self.set_text_color(180, 200, 235)
            self.cell(60, 8,
                      _s(f"{loc_name}  |  {month_year}"),
                      align="R", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        def footer(self):
            self.set_y(-11)
            self.set_fill_color(*BLUE)
            self.rect(0, self.get_y(), 210, 11, "F")
            self.set_font("Helvetica", "", 7)
            self.set_text_color(180, 200, 235)
            self.set_x(10)
            self.cell(0, 11,
                      _s(f"Page {self.page_no()}  |  HPCL SOD MIS  |  CONFIDENTIAL — CHECKER USE ONLY"),
                      new_x=XPos.LMARGIN, new_y=YPos.TOP)

    pdf = _PDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.set_margins(12, 16, 12)
    pdf.add_page()

    W      = 186   # usable width (210 - 12*2)
    LW     = 100   # label column width
    VW     = W - LW  # value column width
    ROW_H  = 5.5   # standard row height
    SUB_H  = 5.0   # sub-section bar height

    # ── Cover card ───────────────────────────────────────────────────────────
    pdf.set_fill_color(*BLUE)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(W, 11, "  MIS REVIEW REPORT", fill=True,
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
    pdf.ln(1)

    pdf.set_fill_color(*LBLUE)
    pdf.set_text_color(*BLUE)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(W, 8, _s(f"  {loc_name}  ({user_id})  |  {zone_name}  |  {month_year}"),
             fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(1)

    pdf.set_text_color(*DARK)
    pdf.set_font("Helvetica", "", 8.5)
    pdf.cell(W, 6, _s(f"  Submission Status:  {_STATUS_LABELS.get(status, status)}"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    # ── Sections S1–S10 ──────────────────────────────────────────────────────
    for sec_num, sec_short in _SECTIONS:
        fields = SECTION_FIELDS.get(sec_num, [])
        if not fields:
            continue

        # Check if this section has any data
        has_data = any(draft.get(f["key"]) not in (None, "", "None") for f in fields
                       if not f.get("auto"))

        # Section header bar
        pdf.set_fill_color(*BLUE)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 10)
        sec_full = SECTION_NAMES.get(sec_num, sec_short)
        pdf.cell(W, 8, _s(f"  Section {sec_num}  —  {sec_full}"),
                 fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(0.5)

        if not has_data:
            pdf.set_text_color(*MGREY)
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(W, 5, "  (no data entered)",
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(2)
            continue

        current_sub = None
        for f in fields:
            sub = f.get("sub", "")
            if sub != current_sub:
                current_sub = sub
                if sub:
                    pdf.set_fill_color(*LBLUE)
                    pdf.set_text_color(*BLUE)
                    pdf.set_font("Helvetica", "B", 7.5)
                    pdf.cell(W, SUB_H, _s(f"  {sub}"),
                             fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                    pdf.ln(0.3)

            key   = f["key"]
            label = f["label"]
            val   = _val(draft.get(key))
            is_auto = bool(f.get("auto"))
            is_text = f.get("type") == "textarea"

            pdf.set_text_color(*DARK)

            if is_text:
                # Label bar (full width, bold, light grey)
                pdf.set_font("Helvetica", "B", 7.5)
                pdf.set_fill_color(*LGREY)
                pdf.cell(W, 4.5, _s(f"  {label}"),
                         fill=True, border="T",
                         new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                # Value — may wrap
                pdf.set_font("Helvetica", "", 8.5)
                pdf.set_fill_color(*WHITE)
                pdf.multi_cell(W, 4.8, _s(f"  {val}"),
                               fill=True, border="B",
                               new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            else:
                row_fill = GREEN if is_auto else WHITE
                pdf.set_fill_color(*row_fill)
                pdf.set_font("Helvetica", "B", 7.5)
                pdf.cell(LW, ROW_H, _s(f"  {label}"),
                         fill=True, border="T",
                         new_x=XPos.RIGHT, new_y=YPos.TOP)
                pdf.set_font("Helvetica", "", 8.5)
                pdf.set_text_color(*BLUE if is_auto else DARK)
                pdf.cell(VW, ROW_H, _s(f"  {val}"),
                         fill=True, border="TL",
                         new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        pdf.ln(3)

    # ── S5A M&I MIS summary ──────────────────────────────────────────────────
    _MI_TABS = [
        ("MI_TANK_OUTAGE",    "S5A-1  Tank Outage",
         ["tank_no","outage_for","planned_start","planned_end","actual_start","actual_end","current_status"]),
        ("MI_MAJOR_REPAIR",   "S5A-2  Major Repair",
         ["tank_no","nature_of_repair","etc_date","current_status"]),
        ("MI_VRU",            "S5A-3  VRU",
         ["vru_no","model","date_not_operating","current_status","remarks"]),
        ("MI_AUDIT_2526",     "S5A-4  M&I Audit 25-26",
         ["audit_no","title","audit_date","status","remarks"]),
        ("MI_AUDIT_2627",     "S5A-5  M&I Audit 26-27",
         ["audit_no","title","audit_date","status","remarks"]),
        ("MI_TECH_AUDIT",     "S5A-6  Tech. Audit",
         ["obs_no","title","audit_date","status","remarks"]),
        ("MI_EQUIP_BREAKDOWN","S5A-7  Equip. Breakdown",
         ["equipment","nature","start_date","proposed_date","actual_end_date","status"]),
        ("MI_INT_PIPELINE",   "S5A-8  Int. Pipeline",
         ["pipeline_no","last_ut_date","last_hydrotest_date","status"]),
        ("MI_EXT_PIPELINE",   "S5A-9  Ext. Pipeline",
         ["pipeline_no","last_ut_date","last_hydrotest_date","status"]),
        ("MI_TANK_STATUS",    "S5A-10 Tank Status",
         ["tank_no","tank_status","cleaning_completed_date","cleaning_due_date","inspection_date"]),
    ]

    any_mi = False
    for tab_key, tab_name, disp_keys in _MI_TABS:
        try:
            rows = load_mi_data(tab_key, user_id, month_year)
        except Exception:
            rows = []
        if not rows:
            continue
        is_na = (len(rows) == 1 and rows[0].get("na_flag") == "Y")
        if not any_mi:
            pdf.set_fill_color(*BLUE)
            pdf.set_text_color(*WHITE)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(W, 8, _s("  Section 5A  —  Maintenance & Inspection (M&I) MIS"),
                     fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(0.5)
            any_mi = True

        # Tab sub-header
        pdf.set_fill_color(*LBLUE)
        pdf.set_text_color(*BLUE)
        pdf.set_font("Helvetica", "B", 7.5)
        suffix = "  —  Not Applicable" if is_na else \
                 f"  —  {len(rows)} row{'s' if len(rows) > 1 else ''}"
        pdf.cell(W, SUB_H, _s(f"  {tab_name}{suffix}"),
                 fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(0.3)

        if is_na:
            continue

        # Determine columns present in data (filter to disp_keys)
        avail_keys = [k for k in disp_keys if any(r.get(k) for r in rows)]
        if not avail_keys:
            avail_keys = disp_keys
        n_cols = len(avail_keys)
        col_w  = min(W / n_cols, 38)

        # Header row
        pdf.set_fill_color(40, 70, 160)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 6.5)
        pdf.set_x(12)
        for k in avail_keys:
            pdf.cell(col_w, 5, _s(k.replace("_", " ").title()[:18]),
                     fill=True, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.ln()

        # Data rows (cap at 30)
        pdf.set_text_color(*DARK)
        pdf.set_font("Helvetica", "", 6.5)
        for ri, row in enumerate(rows[:30]):
            fill_c = LGREY if ri % 2 == 0 else WHITE
            pdf.set_fill_color(*fill_c)
            pdf.set_x(12)
            for k in avail_keys:
                pdf.cell(col_w, 4.5, _s(str(row.get(k, ""))[:22]),
                         fill=True, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP)
            pdf.ln()
        pdf.ln(2)

    return bytes(pdf.output())


# Shared by generate_mi_mis_report() (per-location) and
# generate_mi_mis_consolidated_excel() (all-locations) so both stay in sync.
# Each entry: (tab_key, sheet_name, display_headers, data_keys)
_MI_SHEET_DEFS = [
    ("MI_TANK_OUTAGE", "Tank Outage",
     ["Tank No.", "Other Tank", "Planned Start", "Planned End",
      "Actual Start", "Actual End", "Outage For", "Current Status"],
     ["tank_no", "other_tank_desc", "planned_start", "planned_end",
      "actual_start", "actual_end", "outage_for", "current_status"]),

    ("MI_MAJOR_REPAIR", "Major Repair",
     ["Tank No.", "Other Tank", "Nature of Repair",
      "Revenue/Capex", "AR Code", "Status", "ETC Date"],
     ["tank_no", "other_tank_desc", "nature_of_repair",
      "revenue_capex", "ar_code", "current_status", "etc_date"]),

    ("MI_VRU", "VRU",
     ["VRU Operational", "Date Not Operating", "Action Taken", "ETC Date",
      "MS Vol Recovered (KL)", "Inlet MFM Start", "Inlet MFM End",
      "Outlet MFM Start", "Outlet MFM End", "Vapour Treated (m³)",
      "VOC Value (mg/cc)", "Inlet Emission (mg/cc)",
      "MS/Gasohol TT Vol", "HSD TT Vol", "MS/Gasohol TW Vol",
      "HSD TW Vol", "VRU Uptime %"],
     ["vru_operational", "date_not_operating", "action_taken", "etc_date",
      "ms_vol_recovered_kl", "inlet_mfm_start_m3", "inlet_mfm_end_m3",
      "outlet_mfm_start_m3", "outlet_mfm_end_m3", "vapour_treated_m3",
      "voc_value_mgcc", "inlet_emission_mgcc",
      "ms_gasohol_tt_vol_kl", "hsd_tt_vol_kl",
      "ms_gasohol_tw_vol_kl", "hsd_tw_vol_kl", "vru_uptime_pct"]),

    ("MI_AUDIT_2526", "Audit 25-26",
     ["Audit Date", "No. of Recommendations", "No. Pending", "External Score"],
     ["audit_date", "no_recommendations", "no_pending", "external_score"]),

    ("MI_AUDIT_2627", "Audit 26-27",
     ["Audit Carried Out", "Audit Date", "No. of Recommendations",
      "No. Pending", "External Score"],
     ["audit_carried_out", "audit_date", "no_recommendations",
      "no_pending", "external_score"]),

    ("MI_TECH_AUDIT", "Tech. Audit",
     ["Audit Date", "No. of Recommendations", "No. Pending", "Ref. No."],
     ["audit_date", "no_recommendations", "no_pending", "ref_no"]),

    ("MI_EQUIP_BREAKDOWN", "Equip. Breakdown",
     ["Equipment Name", "Equipment Other", "Equipment Details",
      "Start Date", "Issue", "Proposed Date", "Actual End Date",
      "Resolution Action"],
     ["equipment_name", "equipment_name_other", "equipment_details",
      "start_date", "issue", "proposed_date", "actual_end_date",
      "resolution_action"]),

    ("MI_INT_PIPELINE", "Int. Pipeline",
     ["Last UT Date", "Last Hydrotest Date", "Last DCVG Date",
      "Last LRUT Date", "Other Testing"],
     ["last_ut_date", "last_hydrotest_date", "last_dcvg_date",
      "last_lrut_date", "other_testing"]),

    ("MI_EXT_PIPELINE", "Ext. Pipeline",
     ["Type", "Pipeline Details", "Length (m)", "Product", "Size (inch)",
      "Last UT Date", "Last Hydrotest Date", "Last DCVG Date",
      "Last LRUT Date", "Other Testing"],
     ["pipeline_type", "pipeline_details", "length_metres", "product", "size_inch",
      "last_ut_date", "last_hydrotest_date", "last_dcvg_date",
      "last_lrut_date", "other_testing"]),

    ("MI_TANK_STATUS", "Tank Status",
     ["Zone", "Location", "Tank No.", "Cleaning Completed Date",
      "Cleaning Due Date", "Extension Taken", "eFN No.",
      "Inspection Date", "Inspection Due Date",
      "Painting Date", "Painting Due Date",
      "Tank Status", "Tank Status (Others)"],
     ["zone", "loc_name", "tank_no",
      "cleaning_completed_date", "cleaning_due_date",
      "extension_taken", "extension_efn_no",
      "inspection_date", "inspection_due_date",
      "painting_date", "painting_due_date",
      "tank_status", "tank_status_other"]),
]


def generate_mi_mis_report(
    user_id: str,
    month_year: str,
    user_info: dict,
) -> bytes:
    """Generate filled M&I MIS Excel report with all 10 subsection tabs."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    _thin   = Side(style="thin", color="CCCCCC")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _fill(h):
        return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)

    BLUE_FILL  = _fill("1a1a6e")
    HDR_FONT   = _font(bold=True, color="FFFFFF", size=9)
    NM_FONT    = _font(size=10)
    HINT_FILL  = _fill("F8F9FA")
    HT_FONT    = _font(italic=True, color="888888", size=8)
    WHITE_FILL = _fill("FFFFFF")
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    LOCKED     = Protection(locked=True)

    def _hdr_row(ws, headers, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = HDR_FONT; c.fill = BLUE_FILL
            c.alignment = CENTER; c.border = _border
            c.protection = LOCKED
            ws.column_dimensions[get_column_letter(ci)].width = 22
        ws.row_dimensions[row].height = 28

    def _data_row(ws, row, values):
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = NM_FONT; c.fill = WHITE_FILL
            c.alignment = LEFT; c.border = _border
            c.protection = LOCKED
        ws.row_dimensions[row].height = 18

    def _sheet(tab_key: str, sheet_name: str, display_headers: list, data_keys: list):
        ws = wb.create_sheet(sheet_name)
        rows = load_mi_data(tab_key, user_id, month_year)
        _hdr_row(ws, display_headers)
        if not rows:
            ws.cell(row=2, column=1, value="No data saved for this tab.").font = HT_FONT
        elif rows and rows[0].get("na_flag") == "Y":
            ws.cell(row=2, column=1, value="Not Applicable (marked NA)").font = HT_FONT
        else:
            for ri, rec in enumerate(rows, 2):
                vals = [rec.get(k, "") for k in data_keys]
                _data_row(ws, ri, vals)
        ws.freeze_panes = ws["A2"]
        ws.protection.sheet   = True
        ws.protection.password = "HPCL@MIS"
        ws.protection.selectLockedCells   = False
        ws.protection.selectUnlockedCells = False

    # ── Cover sheet ────────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Cover"
    cover_data = [
        ("Location", user_info.get("locName", user_id)),
        ("Zone",     user_info.get("zone", "")),
        ("Month",    month_year),
        ("Report",   "M&I MIS — Maintenance & Inspection Monthly Information System"),
    ]
    for ri, (k, v) in enumerate(cover_data, 1):
        ws0.cell(row=ri, column=1, value=k).font  = _font(bold=True, color="0033A0", size=11)
        ws0.cell(row=ri, column=2, value=v).font  = _font(size=11)
        ws0.column_dimensions["A"].width = 18
        ws0.column_dimensions["B"].width = 50

    # ── 10 subsection sheets ───────────────────────────────────────────────
    for tab_key, sheet_name, display_headers, data_keys in _MI_SHEET_DEFS:
        _sheet(tab_key, sheet_name, display_headers, data_keys)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_mi_mis_consolidated_excel(month_year: str, rows: list) -> bytes | None:
    """Build ONE workbook with 10 sheets (S5A-1..S5A-10) — each sheet holds every
    SUBMITTED location's rows for that subsection, stacked together and tagged
    with Zone / Location Code / Location Name columns.

    `rows` is a list of location-status dicts (userId/locName/zone/status), e.g.
    from get_all_status_for_month() or get_submissions_for_locations() — callers
    pass in whatever scope/filter (zone, Admin zone-filter, etc.) is on screen.
    TOP/HMEL locations are skipped since M&I MIS (S5) is not applicable to them.
    Returns None if there is nothing to include.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from form_defs import get_skip_sections

    locations = []
    for r in rows:
        if r.get("status") != "SUBMITTED":
            continue
        uid = r.get("userId", "").strip()
        if not uid:
            continue
        if 5 in get_skip_sections(get_loc_type(uid)):
            continue  # M&I MIS not applicable for TOP/HMEL locations
        locations.append({"userId": uid, "locName": r.get("locName", uid),
                           "zone": r.get("zone", "")})
    if not locations:
        return None
    locations.sort(key=lambda l: l["locName"])

    wb = Workbook()

    _thin   = Side(style="thin", color="CCCCCC")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _fill(h):
        return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)

    BLUE_FILL  = _fill("1a1a6e")
    HDR_FONT   = _font(bold=True, color="FFFFFF", size=9)
    NM_FONT    = _font(size=10)
    HT_FONT    = _font(italic=True, color="888888", size=8)
    WHITE_FILL = _fill("FFFFFF")
    ALT_FILL   = _fill("F5F5F5")
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    LOCKED     = Protection(locked=True)

    def _hdr_row(ws, headers, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = HDR_FONT; c.fill = BLUE_FILL
            c.alignment = CENTER; c.border = _border
            c.protection = LOCKED
            ws.column_dimensions[get_column_letter(ci)].width = 20
        ws.row_dimensions[row].height = 28

    def _data_row(ws, row, values, alt: bool):
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = NM_FONT; c.fill = ALT_FILL if alt else WHITE_FILL
            c.alignment = LEFT; c.border = _border
            c.protection = LOCKED
        ws.row_dimensions[row].height = 18

    # ── Cover sheet ────────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Cover"
    cover_data = [
        ("Report",    "Consolidated M&I MIS — Maintenance & Inspection Monthly Information System"),
        ("Month",     month_year),
        ("Locations", str(len(locations))),
    ]
    for ri, (k, v) in enumerate(cover_data, 1):
        ws0.cell(row=ri, column=1, value=k).font = _font(bold=True, color="0033A0", size=11)
        ws0.cell(row=ri, column=2, value=v).font = _font(size=11)
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 60

    hdr_r = len(cover_data) + 2
    ws0.cell(row=hdr_r, column=1, value="Zone").font = _font(bold=True, color="0033A0", size=10)
    ws0.cell(row=hdr_r, column=2, value="Location").font = _font(bold=True, color="0033A0", size=10)
    for ri, loc in enumerate(locations, start=hdr_r + 1):
        ws0.cell(row=ri, column=1, value=loc["zone"])
        ws0.cell(row=ri, column=2, value=f"{loc['locName']} ({loc['userId']})")

    # ── 10 consolidated subsection sheets ────────────────────────────────────
    lead_headers = ["Zone", "Location Code", "Location Name"]
    for tab_key, sheet_name, display_headers, data_keys in _MI_SHEET_DEFS:
        ws = wb.create_sheet(sheet_name)
        _hdr_row(ws, lead_headers + display_headers)
        ri = 2
        for loc in locations:
            mi_rows = load_mi_data(tab_key, loc["userId"], month_year)
            if not mi_rows:
                continue
            lead_vals = [loc["zone"], loc["userId"], loc["locName"]]
            if mi_rows[0].get("na_flag") == "Y":
                _data_row(ws, ri, lead_vals + ["Not Applicable (marked NA)"] +
                          [""] * (len(display_headers) - 1), ri % 2 == 0)
                ri += 1
                continue
            for rec in mi_rows:
                vals = lead_vals + [rec.get(k, "") for k in data_keys]
                _data_row(ws, ri, vals, ri % 2 == 0)
                ri += 1
        if ri == 2:
            ws.cell(row=2, column=1, value="No data for any submitted location this month.").font = HT_FONT
        ws.freeze_panes = ws["A2"]
        ws.protection.sheet   = True
        ws.protection.password = "HPCL@MIS"
        ws.protection.selectLockedCells   = False
        ws.protection.selectUnlockedCells = False

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_full_mis_consolidated_excel_fy(month_year_rows: list) -> bytes | None:
    """One workbook covering every SUBMITTED location for every FY month: a
    "Full MIS Data" sheet with every flat S1-S10 field (read straight from
    MIS_Submitted, one row per location/month) alongside the 10 M&I (S5A)
    sheets. TOP/HMEL locations are included in the flat-data sheet (M&I is
    not applicable to them, but their other sections are) and excluded only
    from the M&I sheets.
    Returns None if there is nothing to include.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from form_defs import get_skip_sections

    all_month_entries = []   # every SUBMITTED location, all sections
    mi_month_entries   = []   # SUBMITTED + M&I-applicable only (excludes TOP/HMEL)
    for my, rows in month_year_rows:
        all_locs, mi_locs = [], []
        for r in rows:
            if r.get("status") != "SUBMITTED":
                continue
            uid = r.get("userId", "").strip()
            if not uid:
                continue
            loc = {"userId": uid, "locName": r.get("locName", uid), "zone": r.get("zone", "")}
            all_locs.append(loc)
            if 5 not in get_skip_sections(get_loc_type(uid)):
                mi_locs.append(loc)
        all_locs.sort(key=lambda l: l["locName"])
        mi_locs.sort(key=lambda l: l["locName"])
        if all_locs:
            all_month_entries.append((my, all_locs))
        if mi_locs:
            mi_month_entries.append((my, mi_locs))

    if not all_month_entries:
        return None

    wb = Workbook()

    _thin   = Side(style="thin", color="CCCCCC")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _fill(h):
        return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)

    BLUE_FILL  = _fill("1a1a6e")
    HDR_FONT   = _font(bold=True, color="FFFFFF", size=9)
    NM_FONT    = _font(size=10)
    HT_FONT    = _font(italic=True, color="888888", size=8)
    WHITE_FILL = _fill("FFFFFF")
    ALT_FILL   = _fill("F5F5F5")
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    LOCKED     = Protection(locked=True)

    def _hdr_row(ws, headers, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = HDR_FONT; c.fill = BLUE_FILL
            c.alignment = CENTER; c.border = _border
            c.protection = LOCKED
            ws.column_dimensions[get_column_letter(ci)].width = 20
        ws.row_dimensions[row].height = 28

    def _data_row(ws, row, values, alt: bool):
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = NM_FONT; c.fill = ALT_FILL if alt else WHITE_FILL
            c.alignment = LEFT; c.border = _border
            c.protection = LOCKED
        ws.row_dimensions[row].height = 18

    # ── Cover sheet ────────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Cover"
    total_loc_months = sum(len(locs) for _, locs in all_month_entries)
    cover_data = [
        ("Report",  "Consolidated MIS (Full + M&I) — All Months, All Locations"),
        ("Months",  ", ".join(my for my, _ in all_month_entries)),
        ("Location-Months Included", str(total_loc_months)),
    ]
    for ri, (k, v) in enumerate(cover_data, 1):
        ws0.cell(row=ri, column=1, value=k).font = _font(bold=True, color="0033A0", size=11)
        ws0.cell(row=ri, column=2, value=v).font = _font(size=11)
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 60

    hdr_r = len(cover_data) + 2
    ws0.cell(row=hdr_r, column=1, value="Month").font = _font(bold=True, color="0033A0", size=10)
    ws0.cell(row=hdr_r, column=2, value="Zone").font  = _font(bold=True, color="0033A0", size=10)
    ws0.cell(row=hdr_r, column=3, value="Location").font = _font(bold=True, color="0033A0", size=10)
    ri = hdr_r + 1
    for my, locs in all_month_entries:
        for loc in locs:
            ws0.cell(row=ri, column=1, value=my)
            ws0.cell(row=ri, column=2, value=loc["zone"])
            ws0.cell(row=ri, column=3, value=f"{loc['locName']} ({loc['userId']})")
            ri += 1
    ws0.column_dimensions["C"].width = 40

    # ── Full MIS Data sheet (flat S1-S10 fields, all months/locations) ──────
    ws_full = wb.create_sheet("Full MIS Data")
    try:
        sub_ws   = _ws(TABS["MIS_SUBMITTED"])
        sub_vals = _api_call(sub_ws.get_all_values)
    except Exception:
        sub_vals = []

    if len(sub_vals) >= 2:
        headers = sub_vals[0]
        uid_idx = next((i for i, h in enumerate(headers) if h.strip().lower() in ("user id", "user_id")), 0)
        mon_idx = next((i for i, h in enumerate(headers) if "month" in h.lower()), 3)
        wanted  = {(loc["userId"], my) for my, locs in all_month_entries for loc in locs}
        _hdr_row(ws_full, headers)
        ri = 2
        for row in sub_vals[1:]:
            row_e = (row + [""] * len(headers))[:len(headers)]
            if (row_e[uid_idx].strip(), row_e[mon_idx].strip()) in wanted:
                _data_row(ws_full, ri, row_e, ri % 2 == 0)
                ri += 1
        if ri == 2:
            ws_full.cell(row=2, column=1,
                         value="No submitted data found for this range.").font = HT_FONT
    else:
        ws_full.cell(row=1, column=1, value="MIS_Submitted sheet not found or empty.").font = HT_FONT
    ws_full.freeze_panes = ws_full["A2"]
    ws_full.protection.sheet   = True
    ws_full.protection.password = "HPCL@MIS"
    ws_full.protection.selectLockedCells   = False
    ws_full.protection.selectUnlockedCells = False

    # ── 10 M&I (S5A) subsection sheets ──────────────────────────────────────
    lead_headers = ["Month", "Zone", "Location Code", "Location Name"]
    for tab_key, sheet_name, display_headers, data_keys in _MI_SHEET_DEFS:
        ws  = wb.create_sheet(sheet_name)
        _hdr_row(ws, lead_headers + display_headers)
        idx = _load_mi_tab_index(tab_key)
        ri  = 2
        for my, locs in mi_month_entries:
            for loc in locs:
                mi_rows = idx.get((loc["userId"], my), [])
                if not mi_rows:
                    continue
                lead_vals = [my, loc["zone"], loc["userId"], loc["locName"]]
                if mi_rows[0].get("na_flag") == "Y":
                    _data_row(ws, ri, lead_vals + ["Not Applicable (marked NA)"] +
                              [""] * (len(display_headers) - 1), ri % 2 == 0)
                    ri += 1
                    continue
                for rec in mi_rows:
                    vals = lead_vals + [rec.get(k, "") for k in data_keys]
                    _data_row(ws, ri, vals, ri % 2 == 0)
                    ri += 1
        if ri == 2:
            ws.cell(row=2, column=1, value="No data for any submitted location in this range.").font = HT_FONT
        ws.freeze_panes = ws["A2"]
        ws.protection.sheet   = True
        ws.protection.password = "HPCL@MIS"
        ws.protection.selectLockedCells   = False
        ws.protection.selectUnlockedCells = False

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def scan_missing_mi_data(month_year_rows: list) -> list:
    """Find SUBMITTED (locked) locations whose M&I (S5A) data is actually
    empty despite being approved -- the exact class of bug caused by the
    (now-fixed) silent upload-failure issue. TOP/HMEL locations are skipped
    (M&I not applicable to them).

    `month_year_rows` — same shape as generate_full_mis_consolidated_excel_fy:
    an ordered list of (month_year, rows) tuples.

    Returns a list of {"userId","locName","zone","month_year","missing_tabs"}
    dicts, one per flagged location-month. Each of the 10 M&I tabs is fetched
    exactly once regardless of scope (same pattern as the consolidated report).
    """
    from form_defs import get_skip_sections

    eligible = []   # (my, loc) pairs to check
    for my, rows in month_year_rows:
        for r in rows:
            if r.get("status") != "SUBMITTED":
                continue
            uid = r.get("userId", "").strip()
            if not uid:
                continue
            if 5 in get_skip_sections(get_loc_type(uid)):
                continue
            eligible.append((my, {"userId": uid, "locName": r.get("locName", uid),
                                   "zone": r.get("zone", "")}))

    if not eligible:
        return []

    tab_indexes = {tab_key: _load_mi_tab_index(tab_key) for tab_key, _, _, _ in _MI_SHEET_DEFS}

    flagged = []
    for my, loc in eligible:
        missing = []
        for tab_key, display_name, _, _ in _MI_SHEET_DEFS:
            rows_for_key = tab_indexes[tab_key].get((loc["userId"], my))
            if not rows_for_key:
                missing.append(display_name)
        if missing:
            flagged.append({
                "userId": loc["userId"], "locName": loc["locName"], "zone": loc["zone"],
                "month_year": my, "missing_tabs": missing,
            })

    return flagged


def _load_mi_tab_index(tab_key: str) -> dict:
    """Fetch one M&I tab ONCE (all rows, all users, all months) and index it by
    (user_id, month_year) -> [row_dict, ...].

    load_mi_data() does one full-tab API fetch per call, which is fine for a
    single location+month but doesn't scale across many locations and months —
    this is the bulk equivalent used by multi-month consolidated reports so each
    of the 10 M&I tabs is only ever fetched once, regardless of scope.
    """
    headers  = _MI_TAB_HEADERS[tab_key]
    ws       = _ensure_ws(TABS[tab_key], headers)
    all_rows = _api_call(ws.get_all_values)
    idx: dict = {}
    if len(all_rows) < 2:
        return idx
    hdr = all_rows[0]
    try:
        uid_idx = hdr.index("user_id")
        mon_idx = hdr.index("month_year")
    except ValueError:
        return idx
    for row in all_rows[1:]:
        row = (row + [""] * len(hdr))[:len(hdr)]
        key = (row[uid_idx].strip(), row[mon_idx].strip())
        idx.setdefault(key, []).append({hdr[i]: row[i] for i in range(len(hdr))})
    return idx


def parse_mis_upload(file_bytes: bytes) -> dict:
    """Parse an uploaded MIS template (.xlsx) and return structured data.

    Returns:
      {
        "fields":         {field_key: value, ...},   # all 140 MIS fields
        "railway_claims": [{data_key: value}, ...],
        "irr_details":    [{data_key: value}, ...],
        "legal_cases":    [{data_key: value}, ...],
        "errors":         [str, ...],                # non-fatal warnings
      }
    """
    import io as _io
    from datetime import datetime as _dt, date as _date
    from openpyxl import load_workbook
    from form_defs import SECTION_FIELDS

    result = {"fields": {}, "railway_claims": [], "irr_details": [],
              "legal_cases": [], "errors": []}

    def _norm(v):
        """Normalise a cell value: datetime/date → DD/MM/YYYY; else stringify."""
        if v is None:
            return None
        if isinstance(v, (_dt, _date)):
            try:
                return v.strftime("%d/%m/%Y")
            except Exception:
                return str(v)
        sv = str(v).strip()
        # Normalise D/M/YYYY or D/M/YY text strings to zero-padded DD/MM/YYYY
        import re as _re2
        _dm = _re2.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})$', sv)
        if _dm:
            _d, _m, _y = _dm.groups()
            if 1 <= int(_m) <= 12 and 1 <= int(_d) <= 31:
                _y = "20" + _y if len(_y) == 2 else _y
                sv = f"{int(_d):02d}/{int(_m):02d}/{_y}"
        return sv if sv else None

    # Build label → key inverse map (case-insensitive, strip * and [Auto])
    label_map = {}
    for fields in SECTION_FIELDS.values():
        for f in fields:
            clean = f["label"].strip().rstrip(" *").replace(" [Auto]", "").strip().lower()
            label_map[clean] = f["key"]

    try:
        wb = load_workbook(_io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        result["errors"].append(f"Cannot open file: {e}")
        return result

    # ── MIS Data sheet ───────────────────────────────────────────────────
    if "MIS Data" not in wb.sheetnames:
        result["errors"].append("Sheet 'MIS Data' not found in uploaded file.")
    else:
        ws = wb["MIS Data"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            result["errors"].append("MIS Data sheet has fewer than 4 rows — expected header + hint + data.")
        else:
            hdr_row  = [str(v or "").strip().rstrip(" *").replace(" [Auto]", "").strip() for v in rows[1]]
            data_row = rows[3]   # row index 3 = row 4 (data)
            N_ID = 4             # first 4 cols are identity (User ID, Location, Zone, Month-Year)
            for ci, hdr in enumerate(hdr_row):
                if ci < N_ID:
                    continue    # skip identity cols
                key = label_map.get(hdr.lower())
                if key and ci < len(data_row):
                    val = _norm(data_row[ci])
                    if val:
                        result["fields"][key] = val

    # ── Detail sheets ────────────────────────────────────────────────────
    def _parse_detail(sheet_name, tab_key, out_key):
        if sheet_name not in wb.sheetnames:
            result["errors"].append(f"Sheet '{sheet_name}' not found — skipped.")
            return
        ddef      = _DETAIL_DEF[tab_key]
        data_keys = ddef["data_keys"]
        col_hdrs  = ddef["sheet_headers"][ddef["prefix_count"]:]

        # Build header → data_key map
        hdr_to_key = {h.strip().lower(): k for h, k in zip(col_hdrs, data_keys)}

        ws2   = wb[sheet_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        if len(rows2) < 2:
            return

        file_hdrs = [str(v or "").strip().lower() for v in rows2[0]]
        col_map   = {ci: hdr_to_key[h] for ci, h in enumerate(file_hdrs) if h in hdr_to_key}

        for raw_row in rows2[1:]:
            if all(v is None or str(v).strip() == "" for v in raw_row):
                continue  # skip blank rows
            rec = {col_map[ci]: _norm(v)
                   for ci, v in enumerate(raw_row)
                   if ci in col_map and _norm(v)}
            if rec:
                result[out_key].append(rec)

    _parse_detail("Railway Claims", "RAILWAY_CLAIMS", "railway_claims")
    _parse_detail("IRR Details",    "IRR_DETAILS",    "irr_details")
    _parse_detail("Legal Cases",    "LEGAL_CASES",    "legal_cases")

    # ── S5A M&I subsection sheets ────────────────────────────────────────
    # Mapping: (Excel sheet name, tab_key, data_keys list)
    _MI_UPLOAD_DEFS = [
        ("S5A-1 Tank Outage",    "MI_TANK_OUTAGE",
         ["tank_no","other_tank_desc","planned_start","planned_end",
          "actual_start","actual_end","outage_for","current_status"]),
        ("S5A-2 Major Repair",   "MI_MAJOR_REPAIR",
         ["tank_no","other_tank_desc","nature_of_repair",
          "revenue_capex","ar_code","current_status","etc_date"]),
        ("S5A-3 VRU",            "MI_VRU",
         ["vru_operational","date_not_operating","action_taken","etc_date",
          "ms_vol_recovered_kl","inlet_mfm_start_m3","inlet_mfm_end_m3",
          "outlet_mfm_start_m3","outlet_mfm_end_m3","vapour_treated_m3",
          "voc_value_mgcc","inlet_emission_mgcc",
          "ms_gasohol_tt_vol_kl","hsd_tt_vol_kl",
          "ms_gasohol_tw_vol_kl","hsd_tw_vol_kl","vru_uptime_pct"]),
        ("S5A-4 M&I Audit 25-26", "MI_AUDIT_2526",
         ["audit_date","no_recommendations","no_pending","external_score"]),
        ("S5A-5 M&I Audit 26-27", "MI_AUDIT_2627",
         ["audit_carried_out","audit_date","no_recommendations",
          "no_pending","external_score"]),
        ("S5A-6 Tech. Audit",    "MI_TECH_AUDIT",
         ["audit_date","no_recommendations","no_pending","ref_no"]),
        ("S5A-7 Equip. Breakdown","MI_EQUIP_BREAKDOWN",
         ["equipment_name","equipment_other","equipment_details",
          "start_date","issue","proposed_date","actual_end_date","resolution_action"]),
        ("S5A-8 Int. Pipeline",  "MI_INT_PIPELINE",
         ["last_ut_date","last_hydrotest_date","last_dcvg_date",
          "last_lrut_date","other_testing"]),
        ("S5A-9 Ext. Pipeline",  "MI_EXT_PIPELINE",
         ["pipeline_type","pipeline_details","length_metres","product","size_inch",
          "last_ut_date","last_hydrotest_date","last_dcvg_date",
          "last_lrut_date","other_testing"]),
        ("S5A-10 Tank Status",   "MI_TANK_STATUS",
         ["tank_no","cleaning_completed_date","cleaning_due_date",
          "extension_taken","extension_efn_no",
          "inspection_date","inspection_due_date",
          "painting_date","painting_due_date",
          "tank_status","tank_status_other"]),
    ]

    result["mi_tabs"] = {}   # {tab_key: [row_dict, ...] or "NA"}

    import re as _re

    def _clean_h(s: str) -> str:
        """Normalise a header for fuzzy matching: strip dots, parens, slashes,
        unicode superscripts, standalone 'of', then remove all non-alphanumeric."""
        s = s.lower().replace("³", "3").replace("²", "2")
        s = _re.sub(r'\bof\b', '', s)
        return _re.sub(r'[^a-z0-9]', '', s)

    for sheet_name, tab_key, data_keys in _MI_UPLOAD_DEFS:
        if sheet_name not in wb.sheetnames:
            result["errors"].append(f"M&I sheet '{sheet_name}' not found — skipped.")
            continue
        ws_mi   = wb[sheet_name]
        mi_rows = list(ws_mi.iter_rows(values_only=True))
        if len(mi_rows) < 4:
            continue  # banner + header + hint rows; data starts at row 4
        hdr_row  = [str(v or "").strip() for v in mi_rows[1]]  # row 2 = headers
        # Build normalised-label → actual-key lookup (both key form and label form)
        clean_to_key: dict[str, str] = {}
        for dk in data_keys:
            clean_to_key[_clean_h(dk)]                 = dk   # e.g. "norecommendations"
            clean_to_key[_clean_h(dk.replace("_", " "))] = dk  # "no recommendations"
        col_map: dict[int, str] = {}
        for ci, h in enumerate(hdr_row):
            hc = _clean_h(h)
            if hc in clean_to_key:
                col_map[ci] = clean_to_key[hc]

        tab_data = []
        had_data_row = False
        for raw in mi_rows[3:]:   # skip banner(0), header(1), hint(2)
            if all(v is None or str(v).strip() == "" for v in raw):
                continue
            had_data_row = True
            # Detect "Not Applicable" marker row
            first_val = str(raw[0] or "").strip().lower()
            if "not applicable" in first_val or first_val == "na":
                tab_data = "NA"
                break
            rec = {}
            for ci, v in enumerate(raw):
                if ci in col_map:
                    nv = _norm(v)
                    if nv:
                        rec[col_map[ci]] = nv
            if rec:
                tab_data.append(rec)

        # Loud failure: the sheet had real data rows, but NOT ONE of them
        # produced a recognised field — almost always a header-text mismatch
        # (edited column headers, or an outdated/newer template version).
        # Without this, the upload silently reports success while quietly
        # dropping this entire M&I section.
        if had_data_row and tab_data == []:
            result["errors"].append(
                f"⚠️ M&I sheet '{sheet_name}' has data but none of its columns "
                f"were recognised — this section was NOT imported. Please "
                f"re-download the current template and re-enter this section."
            )

        result["mi_tabs"][tab_key] = tab_data

    return result


# ── Phase-8/9: Reports & Email ──────────────────────────────────────────────

def get_all_status_for_month(month_year: str) -> list:
    """Return submission status for every Maker location for a given month."""
    locs = get_all_maker_locations()
    return get_submissions_for_locations(locs, month_year)


def download_submitted_data_excel(month_year: str):
    """Build and return an Excel workbook (bytes) of approved MIS data for month_year.

    Returns None if there are no submitted rows for the requested month.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    try:
        ws_sheet = _ws(TABS["MIS_SUBMITTED"])
        all_vals = ws_sheet.get_all_values()
    except Exception:
        return None

    if len(all_vals) < 2:
        return None

    headers = all_vals[0]
    month_col_idx = None
    for ci, h in enumerate(headers):
        if "month" in h.lower():
            month_col_idx = ci
            break

    if month_col_idx is None:
        return None

    data_rows = [row for row in all_vals[1:] if len(row) > month_col_idx and row[month_col_idx].strip() == month_year]
    if not data_rows:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"MIS {month_year}"

    hdr_fill = PatternFill(fill_type="solid", fgColor="002B8F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")

    for ci, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, row in enumerate(data_rows, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center")

    for ci, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(ci)
        max_len = len(str(headers[ci - 1]))
        for row in data_rows:
            if ci - 1 < len(row):
                max_len = max(max_len, len(str(row[ci - 1])))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def download_pending_list_excel(pending_rows: list, month_year: str) -> bytes:
    """Build and return an Excel workbook (bytes) listing pending/non-submitted locations."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    col_headers = [
        "Location Code", "Location Name", "Zone",
        "Status", "Completion %", "Days Overdue / Remark",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Pending {month_year}"

    hdr_fill = PatternFill(fill_type="solid", fgColor="B71C1C")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    for ci, col_name in enumerate(col_headers, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")

    for ri, rec in enumerate(pending_rows, start=2):
        vals = [
            rec.get("userId", ""),
            rec.get("locName", ""),
            rec.get("zone", ""),
            rec.get("status", ""),
            rec.get("completion_pct", 0),
            rec.get("remark", ""),
        ]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center")

    for ci, col_name in enumerate(col_headers, start=1):
        max_len = len(col_name)
        for rec in pending_rows:
            vals = [rec.get("userId",""), rec.get("locName",""), rec.get("zone",""),
                    rec.get("status",""), str(rec.get("completion_pct",0)), rec.get("remark","")]
            if ci - 1 < len(vals):
                max_len = max(max_len, len(str(vals[ci - 1])))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 45)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Analytics helpers ─────────────────────────────────────────────────────────

def _fy_month_years(fy_year: int) -> list:
    """12 'Mon-YYYY' strings for FY starting April fy_year through March fy_year+1."""
    out = []
    for m in range(4, 13):
        out.append(f"{MONTHS_LONG[m - 1][:3]}-{fy_year}")
    for m in range(1, 4):
        out.append(f"{MONTHS_LONG[m - 1][:3]}-{fy_year + 1}")
    return out


_AN_FIELD_LABELS = {
    "f1":  "MS (MT)",
    "f2":  "HSD (MT)",
    "f3":  "Total (MT) incl. Other Products",
    "f4":  "Thruput Target (MT)",
    "f5":  "MEB (Rs in Lakhs)",
    "f6":  "MEB % w.r.t Budget",
    "f7":  "OPEX (Rs/MT)",
    "f8":  "OPEX Target (Rs/MT)",
    "f12": "Electricity Expenditure (Rs in Lakhs)",
    "f21": "SEC (KWH/MT)",
    "f24": "AIM Holds (Nos.)",
    "f26": "Auto-Reconciliation (% of Tanks on Auto Reco)",
    "f38": "CAPEX (Lakhs)",
    "f39": "Capex Target as per AOP (Lakhs)",
    "f46": "MDP Qty Target (KL)",
    "f47": "MDP Qty Actual (KL)",
    "f50": "EBP – Ethanol Blending Percentage",
    "f54": "M&I Index",
    "f55": "PM Percentage",
    "f59": "HSE Index vs Target",
    "f60": "Water Consumed in Month (KL)",
    "f61": "SWC (KL/MT)",
}


@st.cache_data(ttl=300)
def get_compliance_analytics(role: str, zone: str, fy_year: int) -> dict:
    """Fetch submission status for all 12 FY months.

    Returns {month_year: {user_id: {"status", "completion_pct", "submitted_at", "loc"}}}
    """
    months  = _fy_month_years(fy_year)
    locs    = get_all_maker_locations() if role == "Admin" else get_locations_by_zone(zone)
    loc_ids = {l["userId"] for l in locs}
    loc_map  = {l["userId"]: l for l in locs}
    result   = {m: {} for m in months}
    try:
        rows = _submission_status_raw_rows()
        for row in rows[1:]:
            row    = (row + [""] * 9)[:9]
            uid    = row[0].strip()
            mon    = row[1].strip()
            status = row[2].strip()
            pct    = row[3]
            sub_at = row[4].strip()
            if uid in loc_ids and mon in result:
                result[mon][uid] = {
                    "status":         status or "NOT_STARTED",
                    "completion_pct": float(pct) if pct else 0.0,
                    "submitted_at":   sub_at,
                    "loc":            loc_map.get(uid, {}),
                }
    except Exception:
        pass
    for mon in months:
        for uid in loc_ids:
            if uid not in result[mon]:
                result[mon][uid] = {
                    "status":         "NOT_STARTED",
                    "completion_pct": 0.0,
                    "submitted_at":   "",
                    "loc":            loc_map.get(uid, {}),
                }
    return result


@st.cache_data(ttl=300)
def get_analytics_field_data(role: str, zone: str, fy_year: int) -> list:
    """Fetch approved MIS numeric field values for FY from MIS_Submitted.

    Returns list of dicts: user_id, loc_name, zone_name, month_year + field labels as keys.
    """
    months  = set(_fy_month_years(fy_year))
    locs    = get_all_maker_locations() if role == "Admin" else get_locations_by_zone(zone)
    loc_ids = {l["userId"] for l in locs}
    loc_map  = {l["userId"]: l for l in locs}
    needed   = set(_AN_FIELD_LABELS.values())
    try:
        ws       = _ws(TABS["MIS_SUBMITTED"])
        all_vals = ws.get_all_values()
        if len(all_vals) < 2:
            return []
        headers = all_vals[0]
        col_map = {h: ci for ci, h in enumerate(headers) if h in needed}
        uid_c   = next((i for i, h in enumerate(headers) if h in ("User ID", "user_id")), 0)
        mon_c   = next((i for i, h in enumerate(headers) if "month" in h.lower()), 3)
        out     = []
        for row in all_vals[1:]:
            if len(row) <= max(uid_c, mon_c):
                continue
            uid = row[uid_c].strip()
            mon = row[mon_c].strip()
            if uid not in loc_ids or mon not in months:
                continue
            rec = {
                "user_id":    uid,
                "month_year": mon,
                "loc_name":   loc_map.get(uid, {}).get("locName", uid),
                "zone_name":  loc_map.get(uid, {}).get("zone", ""),
            }
            for label, ci in col_map.items():
                raw = row[ci].strip() if ci < len(row) else ""
                try:
                    rec[label] = float(raw) if raw else None
                except ValueError:
                    rec[label] = None
            out.append(rec)
        return out
    except Exception:
        return []


# ── M&I MIS helpers ───────────────────────────────────────────────────────────

_MI_TAB_HEADERS = {
    "MI_TANK_OUTAGE": [
        "user_id", "month_year", "row_no", "na_flag",
        "tank_no", "other_tank_desc",
        "planned_start", "planned_end", "actual_start", "actual_end",
        "outage_for", "current_status", "saved_at",
    ],
    "MI_MAJOR_REPAIR": [
        "user_id", "month_year", "row_no", "na_flag",
        "tank_no", "other_tank_desc",
        "nature_of_repair", "revenue_capex", "ar_code",
        "current_status", "etc_date", "saved_at",
    ],
    "MI_VRU": [
        "user_id", "month_year", "na_flag", "vru_operational",
        "date_not_operating", "action_taken", "etc_date",
        "ms_vol_recovered_kl",
        "inlet_mfm_start_m3", "inlet_mfm_end_m3",
        "outlet_mfm_start_m3", "outlet_mfm_end_m3",
        "vapour_treated_m3", "voc_value_mgcc", "inlet_emission_mgcc",
        "ms_gasohol_tt_vol_kl", "hsd_tt_vol_kl",
        "ms_gasohol_tw_vol_kl", "hsd_tw_vol_kl",
        "vru_uptime_pct", "saved_at",
    ],
    "MI_AUDIT_2526": [
        "user_id", "month_year", "na_flag",
        "audit_date", "no_recommendations", "no_pending",
        "external_score", "saved_at",
    ],
    "MI_AUDIT_2627": [
        "user_id", "month_year", "na_flag",
        "audit_carried_out", "audit_date", "no_recommendations",
        "no_pending", "external_score", "saved_at",
    ],
    "MI_TECH_AUDIT": [
        "user_id", "month_year", "row_no", "na_flag",
        "audit_date", "no_recommendations", "no_pending",
        "ref_no", "saved_at",
    ],
    "MI_EQUIP_BREAKDOWN": [
        "user_id", "month_year", "row_no", "na_flag",
        "equipment_name", "equipment_name_other", "equipment_details",
        "start_date", "issue", "proposed_date", "actual_end_date",
        "resolution_action", "saved_at",
    ],
    "MI_INT_PIPELINE": [
        "user_id", "month_year", "na_flag",
        "last_ut_date", "last_hydrotest_date", "last_dcvg_date",
        "last_lrut_date", "other_testing", "saved_at",
    ],
    "MI_EXT_PIPELINE": [
        "user_id", "month_year", "na_flag",
        "pipeline_type", "pipeline_details", "length_metres", "product", "size_inch",
        "last_ut_date", "last_hydrotest_date", "last_dcvg_date",
        "last_lrut_date", "other_testing", "saved_at",
    ],
    "MI_TANK_STATUS": [
        "user_id", "month_year", "row_no", "na_flag",
        "zone", "loc_name", "tank_no",
        "cleaning_completed_date", "cleaning_due_date",
        "extension_taken", "extension_efn_no",
        "inspection_date", "inspection_due_date",
        "painting_date", "painting_due_date",
        "tank_status", "tank_status_other",
        "saved_at",
    ],
}


def ensure_mi_tabs():
    """No-op on Postgres -- the mi_rows/mi_singletons/mi_submodule_status
    tables already exist via the schema, unlike Sheets worksheet tabs which
    needed explicit auto-creation. Kept as a callable no-op for compatibility
    with any existing call site."""
    pass


# tab_key -> which M&I table stores it: repeatable rows vs one record per
# submission. Matches schema.sql's own split (mi_rows vs mi_singletons).
_MI_MULTIROW_TABS = {
    "MI_TANK_OUTAGE", "MI_MAJOR_REPAIR", "MI_TECH_AUDIT",
    "MI_EQUIP_BREAKDOWN", "MI_TANK_STATUS",
}


_MI_ALL_TABS = (
    "MI_TANK_OUTAGE", "MI_MAJOR_REPAIR", "MI_VRU",
    "MI_AUDIT_2526", "MI_AUDIT_2627", "MI_TECH_AUDIT",
    "MI_EQUIP_BREAKDOWN", "MI_INT_PIPELINE", "MI_EXT_PIPELINE",
    "MI_TANK_STATUS",
)


@st.cache_data(ttl=30)
def check_mi_complete(user_id: str, month_year: str) -> bool:
    """Return True if all 10 M&I MIS tabs have at least one saved row for user+month.

    Cached for 30 s to avoid 10 API calls on every dashboard render.
    """
    for tab_key in _MI_ALL_TABS:
        if not load_mi_data(tab_key, user_id, month_year):
            return False
    return True


_TM_HEADERS = [
    "Sr. No.", "Zone", "Location Code", "Location Name", "SAP Loc Code",
    "Tank No.", "String", "Type", "Year of Commissioning", "Age",
    "Safe Capacity (KL)", "Type2", "SAP Tank No.", "Product",
    "Diameter (m)", "Height (m)",
    "Last Tank Cleaning Date", "Tank Cleaning Due Date",
    "Due for Cleaning 2026-27",
    "Last Comprehensive Inspection Date", "Inspection Due Date",
    "Due for Inspection 2026-27",
    "Last Painted Date", "Due for Painting 2026-27",
    "Cleaning Completed Date", "Cleaning Due Date (Current)",
    "Extension Taken (Yes/No/NA)",
    "Inspection Date (Current)", "Inspection Due Date (Current)",
    "Painting Date (Current)", "Painting Due Date (Current)",
    "Tank Status",
]


@st.cache_data(ttl=7200, show_spinner=False)
def _tm_all_rows() -> list:
    """Read TankMaster sheet ONCE — single cache entry shared by all 111 locations.

    Both get_tank_master() and get_full_tank_master_excel() call this instead
    of hitting the API independently.  No parameters → one cache entry → one
    API read per 2 h regardless of how many concurrent users/locations are active.
    Raises RuntimeError on failure so callers can handle gracefully.
    """
    ws   = _ws(TABS["TANK_MASTER"])
    rows = _api_call(ws.get_all_values)
    return rows or []


def get_tank_master() -> dict:
    """Return {location_code: [tank_no, ...]} for the M&I tank-number dropdowns.

    FLAGGED: the current tank_master table (location_code, tank_no only)
    covers exactly this lookup and nothing more. sync_tank_master_to_sheet()
    and get_full_tank_master_excel() below are intentionally left unported
    -- they need the FULL 31-column maintenance dataset the Sheets
    TankMaster tab carries (cleaning/inspection/painting dates, capacity,
    product, age, etc.). Tank inspection/cleaning scheduling is
    safety-relevant data for a petroleum company; extending the schema for
    it deserves its own careful design pass, not 28 columns bolted on
    mid-migration. Both functions still run against Sheets for now via
    the unchanged _tm_all_rows() below.
    """
    try:
        rows = _pg_query("select location_code, tank_no from tank_master order by tank_no")
        result: dict = {}
        for r in rows:
            result.setdefault(r["location_code"], []).append(r["tank_no"])
        if result:
            return result
    except Exception:
        pass

    # Fallback: local Excel
    try:
        import openpyxl
        path = r"D:\SHOAIB\VS CODE PROJECTS\SOD MIS\M&I Separate Block.xlsx"
        wb_xl = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws_xl = wb_xl["Tank Master"]
        result = {}
        for row in ws_xl.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= 12:
                continue
            loc_code = str(row[2]).strip() if row[2] is not None else ""
            tank_no  = str(row[12]).strip() if row[12] is not None else ""
            if not loc_code or loc_code.lower() in ("none", ""):
                continue
            if not tank_no or tank_no.lower() in ("none", ""):
                continue
            result.setdefault(loc_code, [])
            if tank_no not in result[loc_code]:
                result[loc_code].append(tank_no)
        return result
    except Exception:
        return {}


def sync_tank_master_to_sheet() -> dict:
    """Read Tank Master from local Excel, map zones from UserAccess, write to Google Sheet.

    Returns {"ok": bool, "rows": int, "msg": str}
    """
    try:
        import openpyxl
        from datetime import datetime as _dt

        # Build zone lookup: loc_code_str → full_zone_name from UserAccess
        zone_map: dict = {}
        try:
            ua_ws  = _ws(TABS["USER_ACCESS"])
            ua_rows = _api_call(ua_ws.get_all_values)
            for row in ua_rows[1:]:
                row = (row + [""] * 6)[:6]
                loc_code, _loc_name, zone = row[0].strip(), row[1].strip(), row[2].strip()
                if loc_code and zone:
                    zone_map[loc_code] = zone
        except Exception:
            pass

        # Read Excel Tank Master
        path  = r"D:\SHOAIB\VS CODE PROJECTS\SOD MIS\M&I Separate Block.xlsx"
        wb_xl = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws_xl = wb_xl["Tank Master"]

        data_rows = []
        for ri, row in enumerate(ws_xl.iter_rows(min_row=2, values_only=True), 1):
            if not row or all(v is None for v in row):
                continue
            # Convert row to list; stringify dates and numbers
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                elif isinstance(v, _dt):
                    cells.append(v.strftime("%d/%m/%Y"))
                else:
                    cells.append(str(v))

            # Pad / truncate to 32 columns
            while len(cells) < 32:
                cells.append("")
            cells = cells[:32]

            # Replace abbreviated zone (index 1) with full name from UserAccess
            loc_code_str = cells[2].strip()
            if loc_code_str and loc_code_str in zone_map:
                cells[1] = zone_map[loc_code_str]

            data_rows.append(cells)

        # Write to Google Sheet (clear + rewrite)
        tm_ws = _ensure_ws(TABS["TANK_MASTER"], _TM_HEADERS)
        all_existing = tm_ws.get_all_values()
        if len(all_existing) > 1:
            tm_ws.delete_rows(2, len(all_existing) - 1)

        # Batch write (500 rows at a time) — RAW keeps date strings as-is
        # so Google Sheets does not auto-convert DD/MM/YYYY to date serials.
        BATCH = 500
        for start in range(0, len(data_rows), BATCH):
            tm_ws.append_rows(data_rows[start:start + BATCH], value_input_option="RAW")

        _tm_all_rows.clear()          # clears shared row cache
        get_tank_master.clear()       # clears derived dict cache
        get_full_tank_master_excel.clear()   # clears per-location Excel cache
        return {"ok": True, "rows": len(data_rows),
                "msg": f"Tank Master synced: {len(data_rows)} rows written to Google Sheet."}
    except Exception as exc:
        return {"ok": False, "rows": 0, "msg": str(exc)}


_TM_DATE_COLS = {
    "Last Tank Cleaning Date", "Tank Cleaning Due Date",
    "Last Comprehensive Inspection Date", "Inspection Due Date",
    "Last Painted Date",
    "Cleaning Completed Date", "Cleaning Due Date (Current)",
    "Inspection Date (Current)", "Inspection Due Date (Current)",
    "Painting Date (Current)", "Painting Due Date (Current)",
}


def _normalize_tm_date(val) -> str:
    """Normalise any common date representation → DD/MM/YYYY string.

    Handles:
      DD/MM/YYYY, D/M/YYYY          (already correct or single-digit parts)
      DD.MM.YYYY, DD.MM.YY          (dot-separated)
      DD-MM-YYYY, DD-MM-YY          (dash-separated, day-first)
      YYYY-MM-DD                    (ISO / Google Sheets auto-format)
      YYYY-MM-DD HH:MM:SS           (ISO with time component)
      datetime / date objects       (from openpyxl data_only reads)
    All other values are returned unchanged.
    """
    import re as _re
    from datetime import datetime as _dt, date as _date

    if val is None:
        return ""
    # datetime / date objects (openpyxl returns these for date cells)
    if isinstance(val, (_dt, _date)):
        return val.strftime("%d/%m/%Y")

    if not isinstance(val, str):
        return str(val)

    v = val.strip()
    if not v:
        return v

    # Strip time component from ISO datetime strings
    v_date = v.split(" ")[0].split("T")[0]

    # YYYY-MM-DD  (Google Sheets ISO / standard ISO)
    m = _re.fullmatch(r'(\d{4})-(\d{1,2})-(\d{1,2})', v_date)
    if m:
        y, mo, d = m.groups()
        return f"{d.zfill(2)}/{mo.zfill(2)}/{y}"

    # DD/MM/YYYY or D/M/YYYY
    m = _re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{4})', v)
    if m:
        d, mo, y = m.groups()
        return f"{d.zfill(2)}/{mo.zfill(2)}/{y}"

    # DD/MM/YY or D/M/YY
    m = _re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{2})', v)
    if m:
        d, mo, y = m.groups()
        return f"{d.zfill(2)}/{mo.zfill(2)}/20{y}"

    # DD.MM.YYYY or D.M.YYYY
    m = _re.fullmatch(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', v)
    if m:
        d, mo, y = m.groups()
        return f"{d.zfill(2)}/{mo.zfill(2)}/{y}"

    # DD.MM.YY or D.M.YY
    m = _re.fullmatch(r'(\d{1,2})\.(\d{1,2})\.(\d{2})', v)
    if m:
        d, mo, y = m.groups()
        return f"{d.zfill(2)}/{mo.zfill(2)}/20{y}"

    # DD-MM-YYYY or D-M-YYYY  (day-first dash)
    m = _re.fullmatch(r'(\d{1,2})-(\d{1,2})-(\d{4})', v)
    if m:
        d, mo, y = m.groups()
        return f"{d.zfill(2)}/{mo.zfill(2)}/{y}"

    # DD-MM-YY or D-M-YY
    m = _re.fullmatch(r'(\d{1,2})-(\d{1,2})-(\d{2})', v)
    if m:
        d, mo, y = m.groups()
        return f"{d.zfill(2)}/{mo.zfill(2)}/20{y}"

    return val  # unrecognised — return as-is


@st.cache_data(ttl=7200, show_spinner=False)
def get_full_tank_master_excel(
    location_code: str | None = None,
    zone: str | None = None,
) -> bytes:
    """Return xlsx bytes for Tank Master filtered by location_code, zone, or all rows."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    try:
        all_rows = _tm_all_rows()        # shared cache — no extra API call
    except Exception as exc:
        raise ValueError(f"Cannot read Tank Master: {exc}") from exc

    if len(all_rows) < 2:
        raise ValueError("Tank Master sheet has no data rows.")

    hdr  = all_rows[0]
    data = all_rows[1:]

    try:
        loc_idx = hdr.index("Location Code")
    except ValueError:
        loc_idx = 2
    try:
        zone_idx = hdr.index("Zone")
    except ValueError:
        zone_idx = 1

    if location_code:
        data = [r for r in data if (r + [""] * (loc_idx + 1))[loc_idx].strip() == location_code]
    elif zone:
        data = [r for r in data if (r + [""] * (zone_idx + 1))[zone_idx].strip() == zone]

    # Identify which column indices are date columns
    date_col_indices = {ci for ci, h in enumerate(hdr) if h in _TM_DATE_COLS}

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Tank Master"

    BLUE = PatternFill("solid", fgColor="0033A0")
    W    = Font(bold=True, color="FFFFFF", size=10)
    NM   = Font(size=10)
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    n_cols = len(hdr)
    for ci, h in enumerate(hdr, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = W; c.fill = BLUE; c.alignment = CTR

    for ri, row in enumerate(data, 2):
        row_p = (row + [""] * n_cols)[:n_cols]
        for ci, val in enumerate(row_p, 1):
            if (ci - 1) in date_col_indices:
                val = _normalize_tm_date(val)
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font = NM; c.alignment = LFT

    for ci, h in enumerate(hdr, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = min(len(str(h)) + 4, 35)

    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 30

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_approved_mis_excel(
    zone: str | None = None,
    month_year: str | None = None,
) -> bytes:
    """Return xlsx bytes from MIS_SUBMITTED optionally filtered by zone and/or month."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    try:
        ws_sheet = _ws(TABS["MIS_SUBMITTED"])
        all_rows = _api_call(ws_sheet.get_all_values)
    except Exception as exc:
        raise ValueError(f"Cannot read MIS_SUBMITTED: {exc}") from exc

    if len(all_rows) < 2:
        raise ValueError("No approved MIS submissions found.")

    hdr  = all_rows[0]
    data = all_rows[1:]

    try:
        zone_idx = hdr.index("Zone")
    except ValueError:
        zone_idx = 2
    try:
        mon_idx = hdr.index("Month-Year")
    except ValueError:
        mon_idx = 3

    if zone:
        data = [r for r in data if (r + [""] * (zone_idx + 1))[zone_idx].strip() == zone]
    if month_year:
        data = [r for r in data if (r + [""] * (mon_idx + 1))[mon_idx].strip() == month_year]

    if not data:
        raise ValueError("No approved MIS records found for the selected filters.")

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Approved MIS"

    BLUE = PatternFill("solid", fgColor="0033A0")
    W    = Font(bold=True, color="FFFFFF", size=10)
    NM   = Font(size=10)
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    n_cols = len(hdr)
    for ci, h in enumerate(hdr, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = W; c.fill = BLUE; c.alignment = CTR

    for ri, row in enumerate(data, 2):
        row_p = (row + [""] * n_cols)[:n_cols]
        for ci, val in enumerate(row_p, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font = NM; c.alignment = LFT

    for ci, h in enumerate(hdr, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = min(len(str(h)) + 4, 40)

    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 30

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@st.cache_data(ttl=300, show_spinner=False)
def load_mi_data(tab_key: str, user_id: str, month_year: str) -> list:
    """Return list of row-dicts for user+month from an M&I tab.

    Postgres note: multi-row tabs live in mi_rows, singleton tabs in
    mi_singletons (see _MI_MULTIROW_TABS) -- matching the split the schema
    already has. na_flag is stored once per (submission, tab) in
    mi_submodule_status rather than repeated on every row like the sheet
    did; merged back into each returned dict so the contract (a dict with
    every header key, na_flag included) is unchanged for callers.
    """
    try:
        headers = _MI_TAB_HEADERS[tab_key]
        sub = _pg_one(
            "select id from monthly_submissions where location_code = %s and month_year = %s",
            (user_id, _mk_to_date(month_year)),
        )
        if not sub:
            return []

        na = _pg_one(
            "select is_not_applicable from mi_submodule_status "
            "where submission_id = %s and tab_key = %s",
            (sub["id"], tab_key),
        )
        na_flag = "TRUE" if na and na["is_not_applicable"] else "FALSE"

        if tab_key in _MI_MULTIROW_TABS:
            rows = _pg_query(
                "select row_data from mi_rows where submission_id = %s and tab_key = %s "
                "order by sort_order",
                (sub["id"], tab_key),
            )
            result = []
            for r in rows:
                d = {h: r["row_data"].get(h, "") for h in headers}
                d["na_flag"] = na_flag
                result.append(d)
            return result
        else:
            row = _pg_one(
                "select data from mi_singletons where submission_id = %s and tab_key = %s",
                (sub["id"], tab_key),
            )
            if not row:
                return []
            d = {h: row["data"].get(h, "") for h in headers}
            d["na_flag"] = na_flag
            return [d]
    except Exception:
        return []


def save_mi_data(tab_key: str, user_id: str, month_year: str, rows: list) -> dict:
    """Replace all rows for user+month in an M&I tab with the provided rows list."""
    try:
        headers = _MI_TAB_HEADERS[tab_key]
        sub = _pg_one(
            """
            insert into monthly_submissions (location_code, month_year)
            values (%s, %s)
            on conflict (location_code, month_year) do update set last_updated_at = now()
            returning id
            """,
            (user_id, _mk_to_date(month_year)),
        )
        submission_id = sub["id"]

        na_flag = bool(rows) and str(rows[0].get("na_flag", "")).strip().upper() == "TRUE"
        _pg_query(
            """
            insert into mi_submodule_status (submission_id, tab_key, is_not_applicable, updated_at)
            values (%s, %s, %s, now())
            on conflict (submission_id, tab_key) do update set
                is_not_applicable = excluded.is_not_applicable, updated_at = now()
            """,
            (submission_id, tab_key, na_flag), fetch=False,
        )

        data_cols = [h for h in headers if h not in ("user_id", "month_year", "na_flag", "saved_at")]

        if tab_key in _MI_MULTIROW_TABS:
            _pg_query("delete from mi_rows where submission_id = %s and tab_key = %s",
                      (submission_id, tab_key), fetch=False)
            for sr, rec in enumerate(rows, 1):
                row_data = {c: str(rec.get(c, "") or "") for c in data_cols}
                _pg_query(
                    "insert into mi_rows (submission_id, tab_key, row_data, sort_order) "
                    "values (%s, %s, %s, %s)",
                    (submission_id, tab_key, psycopg2.extras.Json(row_data), sr), fetch=False,
                )
        else:
            rec = rows[0] if rows else {}
            row_data = {c: str(rec.get(c, "") or "") for c in data_cols}
            _pg_query(
                """
                insert into mi_singletons (submission_id, tab_key, data, updated_at)
                values (%s, %s, %s, now())
                on conflict (submission_id, tab_key) do update set
                    data = excluded.data, updated_at = now()
                """,
                (submission_id, tab_key, psycopg2.extras.Json(row_data)), fetch=False,
            )

        audit_log(user_id, f"SaveMI {tab_key}", f"month={month_year} rows={len(rows)}")
        return {"ok": True, "rows": len(rows)}
    except Exception as e:
        return {"ok": False, "msg": str(e)}
